import os, re, time, uuid, threading, json, traceback
import requests as req
from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pptx_generator import generate_pptx, build_campaign_data_from_xlsx

app = Flask(__name__)

UPLOAD_FOLDER = "/tmp/cpt_uploads"
OUTPUT_FOLDER = "/tmp/cpt_outputs"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# ── API KEYS ──────────────────────────────────────────────────────────────────
YOUTUBE_API_KEY   = os.environ.get("YOUTUBE_API_KEY",  "AIzaSyCvfUzTHKJ2m6M6BIMvOrNJeCncaPUlJs8")
REBRANDLY_API_KEY = os.environ.get("REBRANDLY_API_KEY","bf6e85cdd43f4ad68579059ee3dc5882")
TWITCH_CLIENT_ID  = os.environ.get("TWITCH_CLIENT_ID", "sd19zv0wcjyeq4div4qv3unyev4tuh")
TWITCH_SECRET     = os.environ.get("TWITCH_SECRET",    "uaol2jp7pyjlm3bnurma12dhzrcl06")
MODASH_API_KEY    = os.environ.get("MODASH_API_KEY",   "yp58w0tkWbQfsRF7ELMf3cG2NaQpeOV1")

# ── JOB STORE (single dict — fix 404 by using 1 worker in gunicorn) ───────────
jobs = {}
jobs_lock = threading.Lock()

def job_create(jid, data):
    with jobs_lock: jobs[jid] = data

def job_get(jid):
    with jobs_lock: return jobs.get(jid)

def job_update(jid, **kw):
    with jobs_lock:
        if jid in jobs: jobs[jid].update(kw)

def job_log(jid, msg):
    with jobs_lock:
        if jid in jobs: jobs[jid]["log"].append(msg)

def fmt(n):
    if n is None: return "N/A"
    if isinstance(n, str): return n
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000: return f"{n/1_000:.1f}K"
    return str(int(n))

# ── COLUMN AUTO-DETECTION ─────────────────────────────────────────────────────
COLUMN_KEYWORDS = {
    'talent':           ['TALENT'],
    'live_date':        ['LIVE DATE'],
    'content_url':      ['LINK TO CONTENT', 'CONTENT URL', 'CONTENT LINK'],
    'rebrandly':        ['REBRANDLY'],
    'campaign':         ['CAMPAIGN', 'GAME PLAYED', 'GAME'],
    'format':           ['FORMAT'],
    'platform':         ['PLATFORM'],
    'followers':        ['FOLLOWERS'],
    'views':            ['VIEWS'],
    'avg_views':        ['AVG VIEWS', '7 DAY AVG', 'AVG. VIEWS'],
    'avg_ccv':          ['AVG. CCV', 'AVG CCV'],
    'peak_ccv':         ['PEAK CCV'],
    'hours':            ['HOURS WATCHED'],
    'likes':            ['LIKES'],
    'comments':         ['COMMENTS'],
    'shares':           ['SHARES'],
    'saves':            ['SAVES'],
    'total_engagement': ['TOTAL ENGAGEMENT'],
    'engagement_rate':  ['ENGAGEMENT RATE'],
    'link_clicks':      ['LINK CLICKS'],
    'pct_goal':         ['% TO'],
    'budget':           ['BUDGET'],
    'cpc':              ['CPC'],
}

def build_col_map(ws, header_row):
    col_map = {}
    for cell in ws[header_row]:
        if not cell.value: continue
        h = str(cell.value).strip().upper().replace('\n', ' ')
        for key, terms in COLUMN_KEYWORDS.items():
            if key not in col_map and any(t in h for t in terms):
                col_map[key] = cell.column - 1
                break
    return col_map

# ── YOUTUBE ───────────────────────────────────────────────────────────────────
def yt_video_id(url):
    if not url or str(url).strip() in ('', 'N/A', 'None'): return None
    url = str(url).strip()
    for p in [r'youtube\.com/watch\?v=([a-zA-Z0-9_-]{11})',
              r'youtu\.be/([a-zA-Z0-9_-]{11})',
              r'youtube\.com/shorts/([a-zA-Z0-9_-]{11})']:
        m = re.search(p, url)
        if m: return m.group(1)
    return None

def yt_channel_from_url(url):
    """Extract channel ID or handle from a YouTube channel URL"""
    if not url: return None
    url = str(url).strip()
    m = re.search(r'youtube\.com/(?:channel/|@|c/)([a-zA-Z0-9_@-]+)', url)
    return m.group(1) if m else None

def yt_video_stats(video_id):
    try:
        r = req.get("https://www.googleapis.com/youtube/v3/videos",
                    params={"part": "statistics,snippet", "id": video_id,
                            "key": YOUTUBE_API_KEY}, timeout=10)
        items = r.json().get("items", [])
        if not items: return None
        stats   = items[0].get("statistics", {})
        snippet = items[0].get("snippet", {})
        # Get channel subscriber count
        ch_id = snippet.get("channelId")
        subs  = None
        if ch_id:
            cr = req.get("https://www.googleapis.com/youtube/v3/channels",
                         params={"part": "statistics", "id": ch_id,
                                 "key": YOUTUBE_API_KEY}, timeout=10)
            citems = cr.json().get("items", [])
            if citems:
                subs = int(citems[0]["statistics"].get("subscriberCount", 0))
        return {
            "views":     int(stats.get("viewCount", 0)),
            "likes":     int(stats.get("likeCount", 0)),
            "comments":  int(stats.get("commentCount", 0)),
            "followers": subs,
        }
    except: return None

# ── TWITCH ────────────────────────────────────────────────────────────────────
_twitch_token = None
_twitch_token_exp = 0

def get_twitch_token():
    global _twitch_token, _twitch_token_exp
    if _twitch_token and time.time() < _twitch_token_exp:
        return _twitch_token
    try:
        r = req.post("https://id.twitch.tv/oauth2/token",
                     params={"client_id": TWITCH_CLIENT_ID,
                             "client_secret": TWITCH_SECRET,
                             "grant_type": "client_credentials"}, timeout=10)
        data = r.json()
        _twitch_token     = data.get("access_token")
        _twitch_token_exp = time.time() + data.get("expires_in", 3600) - 60
        return _twitch_token
    except: return None

def twitch_handle(url):
    if not url or str(url).strip() in ('', 'N/A', 'None'): return None
    m = re.search(r'twitch\.tv/([a-zA-Z0-9_]+)', str(url))
    if m:
        h = m.group(1)
        # Skip VOD paths
        if h.lower() in ('videos', 'directory', 'clips'): return None
        return h
    return None

def twitch_stats(username):
    """Get Twitch follower count via API + CCV stats via TwitchTracker"""
    result = {"followers": None, "avg_ccv": None, "peak_ccv": None, "hours": None}
    try:
        # 1. Follower count from Twitch API
        token = get_twitch_token()
        if token:
            hdrs = {"Client-ID": TWITCH_CLIENT_ID, "Authorization": f"Bearer {token}"}
            ur = req.get("https://api.twitch.tv/helix/users",
                         params={"login": username.lower()}, headers=hdrs, timeout=10)
            users = ur.json().get("data", [])
            if users:
                uid = users[0]["id"]
                fr = req.get("https://api.twitch.tv/helix/channels/followers",
                             params={"broadcaster_id": uid}, headers=hdrs, timeout=10)
                fdata = fr.json()
                result["followers"] = fdata.get("total")
    except Exception as e:
        pass

    try:
        # 2. CCV stats from TwitchTracker scraping
        hdrs = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                              "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                "Accept-Language": "en-US,en;q=0.9"}
        r = req.get(f"https://twitchtracker.com/{username.lower()}",
                    headers=hdrs, timeout=20)
        if r.status_code == 200:
            page = r.text
            # TwitchTracker embeds stats in JS: {"averageViewers":1234,...}
            m = re.search(r'"averageViewers"\s*:\s*([\d]+)', page)
            if m: result["avg_ccv"] = int(m.group(1))

            m = re.search(r'"peakViewers"\s*:\s*([\d]+)', page)
            if m: result["peak_ccv"] = int(m.group(1))

            m = re.search(r'"hoursWatched"\s*:\s*([\d]+)', page)
            if m: result["hours"] = int(m.group(1))

            # Fallback: parse the stat boxes in the HTML
            if not result["avg_ccv"]:
                # <div class="g-x-s-value">1,234</div> patterns
                vals = re.findall(r'g-x-s-value["\s][^>]*>([\d,]+)<', page)
                labels = re.findall(r'g-x-s-header["\s][^>]*>([^<]+)<', page)
                for lbl, val in zip(labels, vals):
                    lbl = lbl.strip().upper()
                    v = int(val.replace(',',''))
                    if 'AVERAGE' in lbl and 'VIEW' in lbl: result["avg_ccv"] = v
                    elif 'PEAK' in lbl: result["peak_ccv"] = v
                    elif 'HOUR' in lbl: result["hours"] = v
    except Exception as e:
        pass
    return result

# ── REBRANDLY ─────────────────────────────────────────────────────────────────
def rebrandly_clicks(link):
    if not link or str(link).strip() in ('', 'N/A', 'None', '-'): return None
    try:
        link = str(link).strip()
        # Extract slug — handles rebrand.ly/slug AND cherrypick.gg/slug
        m = re.search(r'(?:rebrand\.ly|cherrypick\.gg)/([^/?&#\s]+)', link)
        if not m: return None
        slug = m.group(1)

        api_headers = {"apikey": REBRANDLY_API_KEY, "Content-Type": "application/json"}

        # First try with cherrypick.gg domain
        r = req.get("https://api.rebrandly.com/v1/links",
                    params={"slashtag": slug, "domain.fullName": "cherrypick.gg", "limit": 5},
                    headers=api_headers, timeout=10)
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, list) and data:
                return int(data[0].get("clicks", 0))

        # Then try without domain filter
        r2 = req.get("https://api.rebrandly.com/v1/links",
                     params={"slashtag": slug, "limit": 5},
                     headers=api_headers, timeout=10)
        if r2.status_code == 200:
            data2 = r2.json()
            if isinstance(data2, list) and data2:
                return int(data2[0].get("clicks", 0))

        return None
    except: return None

# ── PROCESS ───────────────────────────────────────────────────────────────────
def process(jid, input_path, output_path, run_yt=True, run_tw=True,
            run_ig=False, run_tt=False, run_rebrandly=True):
    def log(msg): job_log(jid, msg)

    try:
        wb = load_workbook(input_path)
        ws = wb.active

        # Find header row
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row[0] and str(row[0]).upper() == 'TALENT':
                header_row = i
                break

        if not header_row:
            job_update(jid, status="error")
            log("❌ Could not find header row — make sure you're using the CPT template")
            return

        COL = build_col_map(ws, header_row)
        log(f"📋 Detected {len(COL)} columns: {', '.join(list(COL.keys())[:8])}...")

        if 'content_url' not in COL or 'platform' not in COL:
            job_update(jid, status="error")
            log("❌ Required columns missing (LINK TO CONTENT, PLATFORM)")
            return

        # Helper to safely get/set cells
        def get_cell_val(row, key):
            idx = COL.get(key)
            if idx is None: return None
            v = row[idx].value
            return v if v is not None else None

        def set_cell(row, key, val):
            idx = COL.get(key)
            if idx is not None and val is not None:
                row[idx].value = val

        rows = [r for r in ws.iter_rows(min_row=header_row+1, values_only=False)
                if r[0].value and str(r[0].value).strip() not in ('', 'TALENT')]
        total = len(rows)
        log(f"📊 Found {total} creators\n")

        for idx, row in enumerate(rows, 1):
            talent      = get_cell_val(row, 'talent')
            content_url = get_cell_val(row, 'content_url')
            rebrandly   = get_cell_val(row, 'rebrandly')
            platform    = str(get_cell_val(row, 'platform') or '').strip().upper()
            avg_kpi     = get_cell_val(row, 'avg_views')
            budget      = get_cell_val(row, 'budget')

            log(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            log(f"[{idx}/{total}] {talent} ({platform})")
            job_update(jid, progress=int((idx - 1) / total * 85))

            views = likes = comments = 0

            # ── YOUTUBE ──────────────────────────────────────────────────────
            if run_yt and 'YOUTUBE' in platform:
                vid_id = yt_video_id(content_url)
                if vid_id:
                    log(f"   🎬 Fetching YouTube stats...")
                    stats = yt_video_stats(vid_id)
                    if stats:
                        views    = stats['views']
                        likes    = stats['likes']
                        comments = stats['comments']
                        set_cell(row, 'views',     views)
                        set_cell(row, 'likes',     likes)
                        set_cell(row, 'comments',  comments)
                        set_cell(row, 'followers', stats['followers'])

                        total_eng = likes + comments
                        set_cell(row, 'total_engagement', total_eng)

                        # ER as percentage string e.g. "3.25%"
                        if views > 0:
                            er = round(total_eng / views * 100, 2)
                            set_cell(row, 'engagement_rate', f"{er}%")

                        # % to KPI goal
                        try:
                            if avg_kpi and str(avg_kpi) not in ('-','','None'):
                                kpi_f = float(str(avg_kpi).replace(',',''))
                                if kpi_f > 0:
                                    set_cell(row, 'pct_goal',
                                             f"{round(views / kpi_f * 100, 1)}%")
                        except: pass

                        log(f"      ✅ Views: {fmt(views)} | Likes: {fmt(likes)} | "
                            f"Comments: {fmt(comments)} | Followers: {fmt(stats['followers'])}")
                    else:
                        log(f"      ⚠️ Could not fetch YouTube stats for video {vid_id}")
                    time.sleep(0.3)
                else:
                    log(f"      ⚠️ Could not extract video ID from: {content_url}")

            # ── TWITCH ───────────────────────────────────────────────────────
            if run_tw and 'TWITCH' in platform:
                handle = twitch_handle(content_url)
                if not handle:
                    # Fall back to talent name as handle
                    handle = re.sub(r'\s+', '', str(talent or '')).lower()
                if handle:
                    log(f"   🟣 Fetching Twitch stats for {handle}...")
                    stats = twitch_stats(handle)
                    if stats.get('followers'):
                        set_cell(row, 'followers', stats['followers'])
                        log(f"      Followers: {fmt(stats['followers'])}")
                    if stats.get('avg_ccv'):
                        set_cell(row, 'avg_ccv',  stats['avg_ccv'])
                        set_cell(row, 'avg_views', stats['avg_ccv'])
                        log(f"      Avg CCV: {fmt(stats['avg_ccv'])}")
                    if stats.get('peak_ccv'):
                        set_cell(row, 'peak_ccv', stats['peak_ccv'])
                        log(f"      Peak CCV: {fmt(stats['peak_ccv'])}")
                    if stats.get('hours'):
                        set_cell(row, 'hours', stats['hours'])
                        log(f"      Hours Watched: {fmt(stats['hours'])}")
                    if not any(stats.values()):
                        log(f"      ⚠️ No Twitch data found for {handle}")
                    time.sleep(0.5)

            # ── INSTAGRAM (Modash — coming soon) ─────────────────────────────
            if run_ig and 'INSTAGRAM' in platform:
                log(f"   📸 Instagram: Modash API not yet configured")

            # ── TIKTOK (Modash — coming soon) ────────────────────────────────
            if run_tt and 'TIKTOK' in platform:
                log(f"   🎵 TikTok: Modash API not yet configured")

            # ── REBRANDLY ────────────────────────────────────────────────────
            if run_rebrandly and rebrandly and str(rebrandly).strip() not in ('', 'None', '-'):
                log(f"   🔗 Fetching Rebrandly clicks for {rebrandly}...")
                clicks = rebrandly_clicks(str(rebrandly).strip())
                if clicks is not None:
                    set_cell(row, 'link_clicks', clicks)
                    # Add clicks to total engagement
                    cur_eng = get_cell_val(row, 'total_engagement')
                    try:
                        new_eng = int(cur_eng or 0) + clicks
                        set_cell(row, 'total_engagement', new_eng)
                        # Recalculate ER with clicks included
                        v = get_cell_val(row, 'views') or views
                        if v and int(v) > 0:
                            er = round(new_eng / int(v) * 100, 2)
                            set_cell(row, 'engagement_rate', f"{er}%")
                    except: pass
                    # CPC
                    try:
                        if budget and clicks > 0:
                            cpc = round(float(str(budget).replace(',','').replace('$','')) / clicks, 2)
                            set_cell(row, 'cpc', f"${cpc:.2f}")
                    except: pass
                    log(f"      ✅ Link clicks: {fmt(clicks)}")
                else:
                    log(f"      ⚠️ No clicks data — check link format or API key")
                time.sleep(0.2)

        # Highlight auto-filled cells in light olive
        fill = PatternFill("solid", fgColor="E8F0C8")
        for row in ws.iter_rows(min_row=header_row+1):
            if not row[0].value or str(row[0].value).strip() in ('', 'TALENT'):
                continue
            for key in ['views','likes','comments','followers','avg_ccv','peak_ccv',
                        'hours','total_engagement','engagement_rate','link_clicks',
                        'pct_goal','cpc','avg_views']:
                idx = COL.get(key)
                if idx is not None and row[idx].value not in (None, '', '-'):
                    row[idx].fill = fill

        wb.save(output_path)
        log(f"\n📊 Generating EOC PowerPoint deck...")
        try:
            pptx_path = output_path.replace('.xlsx', '.pptx')
            campaign_data = build_campaign_data_from_xlsx(output_path)
            generate_pptx(campaign_data, pptx_path)
            job_update(jid, pptx=pptx_path,
                       pptx_name=os.path.basename(pptx_path))
            log("✅ PowerPoint deck generated!")
        except Exception as pe:
            log(f"⚠️ PPTX generation error: {pe}")

        job_update(jid, status="complete", progress=100)
        log("\n🎉 All done! Download your filled report and EOC deck below.")

    except Exception as e:
        job_update(jid, status="error", progress=0)
        job_log(jid, f"❌ Error: {str(e)}")
        job_log(jid, traceback.format_exc())

# ── ROUTES ────────────────────────────────────────────────────────────────────

@app.route("/api/run", methods=["POST"])
def run_job():
    if "file" not in request.files: return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.endswith(".xlsx"): return jsonify({"error": "Please upload an .xlsx file"}), 400
    run_yt = request.form.get("run_yt", "1") == "1"
    run_tw = request.form.get("run_tw", "1") == "1"
    run_ig = request.form.get("run_ig", "0") == "1"
    run_tt = request.form.get("run_tt", "0") == "1"
    run_rb = request.form.get("run_rb", "1") == "1"
    jid  = str(uuid.uuid4())
    base_name = os.path.splitext(secure_filename(f.filename))[0]
    inp  = os.path.join(UPLOAD_FOLDER, f"{jid}_input.xlsx")
    out  = os.path.join(OUTPUT_FOLDER, f"{jid}_{base_name}_Filled.xlsx")
    f.save(inp)
    job_create(jid, {"status": "running", "progress": 0, "log": [],
                     "output": out, "name": f"{base_name}_Filled.xlsx",
                     "pptx": None, "pptx_name": None, "ts": time.time()})
    t = threading.Thread(target=process,
                         args=(jid, inp, out, run_yt, run_tw, run_ig, run_tt, run_rb),
                         daemon=True)
    t.start()
    return jsonify({"job_id": jid})

@app.route("/api/status/<jid>")
def job_status(jid):
    j = job_get(jid)
    if not j: return jsonify({"error": "Job not found"}), 404
    return jsonify({"status": j["status"], "progress": j["progress"],
                    "log": j["log"], "has_pptx": bool(j.get("pptx"))})

@app.route("/api/download/<jid>")
def download(jid):
    j = job_get(jid)
    if not j: return jsonify({"error": "Job not found"}), 404
    return send_file(j["output"], as_attachment=True, download_name=j["name"])

@app.route("/api/download-pptx/<jid>")
def download_pptx(jid):
    j = job_get(jid)
    if not j: return jsonify({"error": "Job not found"}), 404
    pptx = j.get("pptx")
    if not pptx or not os.path.exists(pptx):
        return jsonify({"error": "PPTX not available"}), 404
    return send_file(pptx, as_attachment=True,
                     download_name=j.get("pptx_name", "CPT_EOC_Deck.pptx"))

@app.route("/")
def index():
    LOGO = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCANHB9ADASIAAhEBAxEB/8QAHQABAAIDAQEBAQAAAAAAAAAAAAgJBQYHBAIDAf/EAGEQAQABAgQCAwYMEgYIBQMEAwABAgMEBQYRByEIEjEJE0FRdbMYIjI3OFZhcXSRlNEUFRcjMzZCU1VXcnOBlbGytNIWNVJ2hdMkQ1RikqHC4WOCpcHUJTTkRISTooOj8P/EABQBAQAAAAAAAAAAAAAAAAAAAAD/xAAUEQEAAAAAAAAAAAAAAAAAAAAA/9oADAMBAAIRAxEAPwCGQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMnl+n8+zCaIwGS5jiuvG9PecLXX1o8cbQDGDJ5hp/PsvmuMfkuY4XqRvV37C10dWPHO8MYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA3rgzws1VxV1FXlOm8NEWsPFNWMxt3eLOGpqmYiap8c7TtTHOdp8U7a1pDT+aar1RlunMlw9WIzDMcRTYsUR45ntmfBERvMz4IiZWqcH9AZPw00JgdLZPTFVNmOvicR1IpqxN6YjrXKvdnbl27RER4AaPwi6NnDfQGFtXr2W29RZxTMVV5hmNuKtqo+92+dNEfHPjmXY8PYs4axRYw9m3ZtURtRRbpimmmPFER2P0AfniLFnE2K7GIs271quOrXRcpiqmqPFMT2uPcXejdw21/hL123ldvIM5q528wy6iKJ63+/b9TXHZ4InlymHZQFUnGfhRqvhVqH6Waiw0VYa9NX0HjrO82cTTE9sT4J7N6Z5xu0JbVxc0BkfErRGN0xnlmKqL1PWw9+nlXh70R6W5TPgmJ+ON4ntVW6007mWktV5npvN7U2sbl2IqsXY25TMTyqj3JjaY9yYBhwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAASp4L9FDLuIPDXKNXXtaYvL7mYW6q5w9GApuRRtVNPqprjfs8QIrCbXoHco/GJjv1ZT/mHoHco/GJjv1ZT/mAhKJtegdyj8YmO/VlP+Yegdyj8YmO/VlP+YCEom16B3KPxiY79WU/5h6B3KPxiY79WU/5gISjKatyunI9VZtktF6b9OX429hYuTT1Zri3XNPW28G+27FgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlD3OvTFrMuJ2bakxGHmuMowPVsVzRE003bs9Xtnsq6sVc48Eynuh/3NGY+kut43jf6Iwc7f+W6mAAAAAAgd3RrS2Fy3iLkWqsLbpt15zgqrWKimjbrXbExEVzPhmaK6KfetwniiV3SiuxGjtJ2pmn6InMLtVMeHqxb5/o3mAQcAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAWedDz2OmlPzFzzlSsNZ50PPY6aU/MXPOVA64AAAAACozip65+qvLOL89W1psvFT1z9VeWcX56trQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAJLdz11VZyfi7jNPYiuaac8wNVFneqdu+2vTxG3jmmK/iWBKc8jzPHZLnWBzjLL9WHx2BxFvE4a7TETNFyiqKqaufLlMQtH6P/FPKeK2g8PnWEqt2cysxFrMsHFcTVYuxHOdu3q1dsT4uXgkHRQAAAEAu6H6uw+dcVct0xhKqLlGQYKYv1RvvF+9MV1UeKYiim1Pv1VR4EuuPfFTI+FOiMRnOY3Iu5hdpm3luBomOviL0xy96iO2qqeyI5bzMRNXOoc3x+f57js7zS/N/HY6/XiL9yfuq6p3mf+YPAAAAAAAAAAAAAAAAAAAAAAAzWjNK6h1lndOS6Yyu9meYVW6rkWLUx1ppp7Z5zEcmFSA6AfshcL5NxX7sA1H0PPGj8X+af8Vv+Y9Dzxo/F/mn/Fb/AJlowCrn0PPGj8X+af8AFb/mPQ88aPxf5p/xW/5lowCrn0PPGj8X+af8Vv8AmPQ88aPxf5p/xW/5lowCn7V+ms90jn17ItSZbey3MrFNNVzD3ZjrUxVTFVPZMxziYliHcenT7JPPvg+E/h7bhwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACzzoeex00p+YuecqVhrPOh57HTSn5i55yoHXAAAAAAVGcVPXP1V5Zxfnq2tNl4qeufqryzi/PVtaAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAbLw41zqbh9qaxqDS+Y14PF2piK6e23fo350XKeyqmfF8W082tALBuEnS70HqS1ZwWsoq0tmdW1M3K4quYSuqZiI2riN6O3f08RTG071O+ZTqbTebYKjHZVqDKcfha9+pfw2Mt3LdW3btVTMxKn0BcHmupdOZTgq8dmmoMqwOFo9XexOMt26KffqqmIhwTi50uNBaZtYjAaPmrVGbUTNEXLUTTg6Kucbzcn1cbxHqImJieVSvcBs/EvXepuImqcRqPVGYVYrFXp2t26Y6trD0eC3bp7KaYj9M9szMzMzrAAAAAAAAAAAAAAAAAAAAAAAAJAdAP2QuF8m4r92Ef0gOgH7IXC+TcV+7ALGAAAAAAVsdOn2SeffB8J/D23DncenT7JPPvg+E/h7bhwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA9eByzMsdG+Cy/F4qN+rvZs1V8/FyhtdjhHxTv2ab1nh1qu5brjemqnKb0xMe56UGkjfK+DPFqinrTw31VMe5ll2Z+KKWnZxlmY5PmN7Lc2wOIwONsztdsYi3NFyifFNM84B5B92LVy/eos2aJruXKopopjtmZnaIb3TwX4tVRExw51PMTziYy65z/wCQNBG//UW4t/i41R+rrnzH1FuLf4uNUfq658wNAG//AFFuLf4uNUfq658x9Rbi3+LjVH6uufMDQBv/ANRbi3+LjVH6uufM0CeU7SAs86HnsdNKfmLnnKlYazzoeex00p+YuecqB1wAAAAAFRnFT1z9VeWcX56trTZeKnrn6q8s4vz1bWgAAAABmMk0tqbPL9uxk2n81zG7dje3ThsJXcmv3to5towfBTi5i7k0WuHGpqZj77l9y1Hx1REA5+OhY3gjxdwc7XeHOpKue31nA13f3Ilq+faU1RkN6uznenc2y25bjeunFYO5bmmPHPWiNgYUAAAAAAAAAAAAAAHtwGU5pmHV+gctxmK61XVjvNiqvefFygHiG84fg/xWxEUTa4c6qqpubTTV9K70UzE+HeadmM1toLV+iqcNOq8jxGUVYrebFvEVUxcriO2Yo3620cue23OAayD+xG8xHjB/BsuZaB1vluX2Mwx2ks8w+CxERNjEV4G53q7ExvE017bVRMeGJa5cort11W7lFVFdM7TTVG0xIPkAAAAAAAAAAAAAAAAAAAAHuynKM2ze/TYyrLMZjrtVUUU0YexVcmap7I5R2g8I6DgeCXF3GVVU2eHOpaZp7e/YGu1Hx1xG7SM2y/GZVmmKyvMcPXhsbg71djEWa/VW7lEzTVTPuxMTE+8DygAAAAAAAAAAAA/sRMztEbyD+DZsh4fa7z+1XdyTRuf5lRbnaurDZfduRTPu7UshqnhPxH0tp+rP9RaQzPK8toqppqv4iiKYpmqdqYmN943n3AaSAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP0w9m9iLsWsPauXblXZRRTNUz+iGx5Hw815nluq5k2jNQZhRRyqqw+XXbkR8VINYHQ8FwP4vYuPrXDrUdPg+vYKq1+/s/mM4I8XcJt33hzqWrf71ga7v7sSDno2HPdDa0yK7TazrSeeZdXVG8U4nAXLczHj5w1+YmJmJiYmO2JB/AAAAAAAAAAAAAAAAAAAAAAAAAABnMj0fqzPcRGHybTOcZjdqp60UYbBXLkzHj5Q2nLeBnGDMKpixw61DRtO3+kYSbHnOqDnQ6fmfR+4yZbh6sRi9BZnFqimaqqrdVu51YjnMz1ap2cyu267V2u1cpmmuiqaaqZ7YmO2AfIyWmskzPUeeYXJMnsUYjH4uvqWLVV6i116v7MVVzEbz4I35zyh0n0NnG72hYv5Xh/8wHJB1v0NnG72hYv5Xh/8w9DZxu9oWL+V4f/ADAckHW/Q2cbvaFi/leH/wAw9DZxu9oWL+V4f/MByQdPzvo/8YMlyjF5tmeisVhsFg7NV7EXZxNiYoopjeZ2iuZ5R4nMAEgOgH7IXC+TcV+7CP6QHQD9kLhfJuK/dgFjAAAAAAK2OnT7JPPvg+E/h7bhzuPTp9knn3wfCfw9tw4AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACImZ2jnIDK6U03n2q85tZNpvKcZmuYXfUWMNbmurbw1T4qY8MztEeFIDo89FbP8AWn0NqDW838i0/VTTdt2IjbFYuJ8ERP2On/emN53jaPDE4dAaI0toPJaco0rk2Gy3DbR15t0+nuzEbRVXXPOqe3nM+GQQ44adC/UeZUWsXrrPrOS2aurVVhMHTF+/Mcpmmap9JTO28b+m2mOyUjtDdHHhBpOxajD6UsZniqIp62LzSqcTcrmJ3irafSUz+RTS62A82X5fgMutTay/A4bCW6p3mmxapoiZ8e0Q9M8o3kQ06afSDxWHxeN4Z6Jxc2Zt72s5zC1V6bfw4e3Mep2+6nt39Ly2ncP70rOlDesYnE6K4Y5hFNVuarWYZza5zv2TbsT+25H/AJfGhnduXLt2q7drquXK5mqqqqd5qme2Zl8gCYPQ96SH0D9CcPuIeYzOEmYtZVmuIq+w78qbN2qfuPBTXPqd9pnqxG0PgFzAiD0LOkF9MKMLw21tjY+jKerayfHXavs0dkWK5n7r+zPh7O2I3l8AAAppufZKvflcsppufZKvfkHys86HnsdNKfmLnnKlYazzoeex00p+YuecqB1wAAAAAFRnFT1z9VeWcX56trTZeKnrn6q8s4vz1bWgGU0tp3PdU51YyXTuVYvNMwvztbsYa3NdXvz/AGaY8NU7REc5mHaujv0Z9ScSJw2e5/VeyHS9cRcovVW/r+Mp37LVM9kTG/1yd47JiKk8eHHD7SPD3JoyvSmTYfAW5iIu3Yp3u3pjw11zzq/T45BEThd0MM8zGizjeIGeU5PZqiKqsDgOrdxG3hpmud6KZ92Iq5wkxoXgNwo0bRZqynR+BvYq1Ef6XjonE3qqop2mreveKZntmKYiOfKIdMAfnhsPYwtijD4azbsWaI2ot26Ypppj3IjlD9AAfnisPYxWHrw+KsWr9muNq7dyiKqao8UxPKX6AOV8Rej5wp1xYu/TDTNnLsbXMzGOyzbD36Znbnyjq1dn3dNURz2RI4zdE3WujrF/NdLXZ1TlVqJrrptW+rirdMRMzM2956+233O8+4sLAU0VU1U1TTVE01RO0xMc4l/Fh/Sg6N2Ua/y/F6k0jhsNluraJm9XTTHUtZh/apr25U3J7Yr8M8qu3rRX1muX43KsyxOW5lhbuExmGuVWr9m7T1ardcTtMTAPKAAAAAAP1weGxGMxVrC4TD3cRiLtUUW7Vqiaq66p7IiI5zPuJdcA+iDisbFrPeKVy5g7HKq1k1iqIu184ne9XHqY23jqU8+fbG20hGHQuitV65zb6V6TyLG5tiY5194t+ktRz511z6WiOU86pjeeXak5w46FWZ4qLOK17qajAW52mvB5ZTFy7tvPLvlUdWmeyd+rV7yY+ltOZFpbKLWUadynCZZgbUR1bOHtxREztEbzt2ztEc53llAcr0R0euEWkrNr6B0hhMbibfVmcXmMzibtVVP3Xp/S0z44oimPcdMy/AYHL7M2cBgsNhLUzvNFi1TRTv70Q9DTuMfEHJuGeg8bqjObkbWo73hbEeqxF+YnqW6Y93aZnxREz4Aar0luNWV8ItMU1UW7WO1DjqKoy/BVVel5cu+3NufUifBHOezeO2K2tY6mzzV+oMTn2osxvY/MMTVvcu3J7I8FMR2RTHgiOUPRxC1hn2u9WYzU2o8ZXisdiqvDPpbVEept0R9zTEdkR7/bMtfAABMToI8bLtnE2eFepsTTVh7kzOSYiuJ61FXbVh5nsmmec0+GJ3jeYmmIlTqvhvoLVVqbeoNI5Pj9+tPXuYWmK4me2YqjaYn3YndUvhMRiMJirOLwl+5YxFmum5au26pprorpneKomOcTExvErP8Aov8AE+jijwwwuZ4m7ROdYLbC5pRTER9diOVe2/KK45+/v4gcw4j9DLRWa98xWis2xuncRNPpcLfqnE4beI5RE1T3yneeczNVXuRCLvFrgJxI4b27uNzfJqsblFv1WZYHe7ZoiZ2jr7c6PBzqiI5xG+60UmImJiY3ie2AUziwvjz0VtJ6zw2IzbRtuxpzP9pqpot09XB4ieXKuiI9JPbtVT4+cSgjrnSWodE6jxOn9TZbewGPw9UxNFcb01x4KqKo5VUz4JgGCAAABaDorg3wpxWjckxOJ4e6bu3r2X4e5cuV4Ciaqqpt0zMzO3bMsv8AUU4R/i40z+r7fzNl0B9omn/JmG81SzYOf/UU4R/i40z+r7fzH1FOEf4uNM/q+38zoADn/wBRThH+LjTP6vt/MfUU4R/i40z+r7fzOgAOX6m4M8J7Gm8zv2eHmm7d23g7tdFdOAoiaZiiZiY5Kt1wurftUzf4De83Up6AB++AwmKx+NsYHBYe7icViLlNqzZtUzVXcrqnaKYiOczMg/B0fg9wW17xQxEV6fyqq1llNcUXczxW9vD0Tz3iKvu5jbnFO+3LfbeEi+jt0SLNu1Y1HxUoi5cq9NZyOir0tMctpv1xPOZ5+kjlttvO+8RL/L8Fg8uwVrA5fhbGEwtmnq2rNm3FFFEeKKY5RAI88J+iNw+0xRaxurJuarzOIpqmm/vbwlurw9W3E718949PMxMbelh3/JcmyjJMLThcnyzB5fYpiKYt4azTbjaI2jsh7gHlzjMMLlGUYzNcdc71hMFh68Rfr236tFFM1VTt7kRKn7P80xmeZ7mGdZjci7jcwxNzFYiuI2iq5cqmqqdvBvMys46WOf8A9Hej/qrF035sXsRhfoO1VFO+9V2qKJj9NM1Qq6B3roM6Z09qvjFi8t1Lk2BzfBU5Peu02MXZi5RFcXLURVtPh2mfjTe+opwj/Fxpn9X2/mQ47nd6+mN8h3/O2Vg4Of8A1FOEf4uNM/q+38x9RThH+LjTP6vt/M6AA5/9RThH+LjTP6vt/MfUU4R/i40z+r7fzOgAOf8A1FOEf4uNM/q+38yKvT80TpHSFvSs6X03lmTTiZv9++g8PTb75t1dt9u3beU6kNe6V/YtG/lYn/oBDEAB+2Dw2JxuLs4PB4e7icTfuU27Nm1RNddyuqdopppjnMzMxERDduDnCfV/FPO5y/TWCj6Hs1U/ReOvT1bGHiZ7ap8M7bz1Y3mdlg/A/gRonhXhLd7L8NGZZ31Ji9muKoibszPb1I5xbjwbRz27ZnmCLPB7ogas1Hbw+Z65xdWmsvrimv6FppirGV0zz2mJ5W5/K3mPDT4Es+HfA7hhoXD2qcl0thL2LooimrG42n6IxFydoiapqq5RvtvMUxTTv2RDpAD+UxFNMU0xEREbREeBFvujmofoDhnkOnLeIrt3s2zKq7Xbinem5ZsUb1bz4Nq7lqUpVf8A3RLUNWZcYsuyG3fu1WMnyuiKrVUbU0Xr1U11VU+PejvMT+SCM6TfQUt8P9R57mmidZ6XyTNMbfp+i8tvYzC26656sbXLUTMbzy2qiN522qnbtRkZjRWosy0lqzK9S5PfqsY7LsTRftVUztvtPOmfHTVG9Mx4YmY8ILQPqKcI/wAXGmf1fb+Y+opwj/Fxpn9X2/mbJoHU+X6y0blWp8rrpqwuY4em9TEVdbqTMemomfHE7xPZ2M4Dn/1FOEf4uNM/q+38x9RThH+LjTP6vt/M6AA5/wDUU4R/i40z+r7fzH1FOEf4uNM/q+38zoACv/pzcHcBojUOC1dpbLLeCyDMqabF/DYe3FNrC4imnaOrERtTTXTETt/aiqfDtEZ1tnF3RGXcReHmbaSzLq0U42xPeL80dacPfjnbuxG8b9Wradt43jeN9plVFqHKMwyDPcdkma4erD47A36rF+3P3NdM7T+j3QeAAAAAAGS0xmVnJ9RYDNMRl2FzOzhb9Ny7g8TRFVq/RE+moqiYmNpjeN9p27VluhuHXA7WWkcs1PknD/TV7AZjYi9amcutxVT4JpqjblVTMTTMeOJVgJYdADitRk+oL3DTOb8UYLNK5vZXcqnlRido61rs+7pjeJ3jnTtz60bBKj6inCP8XGmf1fb+Y+opwj/Fxpn9X2/mdAAc/wDqKcI/xcaZ/V9v5j6inCP8XGmf1fb+Z0ABz/6inCP8XGmf1fb+ZDfpw3+H+TaowWhtD6VyHLL2A/0jNMXgsNRRcm5VTtRZ61PZEUzNVUT2zNPZ1ec1+M+u8Dw34cZrqzG003asLa6uFsTO3f79XK3Rv4Ime2fBETKqfUWcZjqDPcbnebYmvE4/HX6r9+7VPOqqqd596PFHgjkDwAAO8dBvTWn9V8ZsTlmpcmwWbYKnJ792LGLsxcoiuLlqIq2nwxEz8bg6R/c8PX5xXkLEedsgmT9RThH+LjTP6vt/MfUU4R/i40z+r7fzOgAOf/UU4R/i40z+r7fzH1FOEf4uNM/q+38zoADn/wBRThH+LjTP6vt/MfUU4R/i40z+r7fzOgAK2em5pzIdLcbPpXpzKMFlOB+leHufQ+FtRbo60zXvVtHhnaHDUh+6C+yA/wAIw37a0eAdX4QcJsh1nZoxmf8AFTR+l8NMxvZxeNojFc9/9XXVRG/KPD4UwuHfRT4PZJhbGLxeFxOq79UUXKcRjsVvZnl9zbt9WiaJ7dquv76uZmdL6q1LpfFfROnM+zHKrk1RVVOFxFVuK5js60RO1X6dwWxae0dpPT1m1ZyPTeU5dRamZt/Q+EoomnfxTEbs6gFwr6YetshvWMHrXBWdS5dHVprvU7WcXRG8b1RVEdWudt52qiN529NSmPwq4qaJ4mZb9F6Vze3fvUURVfwd30mIs89vTUT4N/DG8dnPmDdgAKoiqJpqiJieUxPhaTrHhNw21dZm3qDRuUYqZpmmLlNjvV2mJnedq6Nqo5+KW7AIecUehZgrvfsdw61Bdw9czNUZdmc9eiO2erRdpjeI7IiKomfDNUoncQdC6s0DnU5RqzJMVlmInnbquU7271PL01uuPS1xzjfaZ2nlO08lujCa10np3WeRXck1NlOGzLA3efe71O80z4KqZ7aZ92AVBCQfSb6N2bcNKruo9N13s20rXXM1TNO9/A7zypubeqo8Vce9MRymY+AAAAA6j0UcmynUHSA0vk+eZdhsxy/EXMRF7DYi3Fdu5thrtUbxPKdpiJ/QsH+opwj/ABcaZ/V9v5kB+hj7JjR/5zE/wt5ZwDn/ANRThH+LjTP6vt/MfUU4R/i40z+r7fzOgAOf/UU4R/i40z+r7fzH1FOEf4uNM/q+38zoADn/ANRThH+LjTP6vt/MiN0/tG6U0hnmkrWltPZbk1vE4bE1X6cHYptxcmmq3tM7du28/GnyhJ3S37YtFfBMX+/bBEIAAAAHoy7BYvMcdYwGAw17FYq/XFuzZtUTVXcqnsiIjnMg87cOGnDPW/EXHzhdJ5BisdRRVFN7FTT1MPZn/fuT6WJ8O2+8+CJSh6P3RDt0UWNQcVJ69zrRXaySzV6WIjaYm9XE89+fpKfBtvVzmIl1kmU5ZkmV2MryfAYbAYKxT1bVjD24oopj3IgESuG3QqwFuLWL4galv36toqnA5XtRTE8p2qu1RMzHbExTEeOKoSH0fwd4Y6St005ForKLFVNM099u2e/3ZiZ32mu51qp/TLewHzat27Vum3aopoopjammmNoiPFEPNnOZYDJsqxWa5ri7ODwOEtVXsRfu1dWi3RTG8zMv3xN+zhsPdxOIu0WbNqia7lyuqKaaKYjeZmZ7IiPCrp6WnHrHcSs+v6cyDE3LGkMFe2t00TMfR9dM/Za/HTv6mmfFEzz7A9HSd6SOccRMVitN6Wu3ss0lTV1Ktp6t7H7T6q5/Zo8VH6Z37IjyAPq3XXauU3LddVFdExVTVTO0xMdkxKe/Q36QMaywWH0JrLHUzqPD0TTgsXdmInH26Y36s+O7TEe/VEb853mYDP2wOKxOBxljG4O/cw+JsXKblq7bq6tVFdM7xVEx2TEwC5IR26JfSEwvEbL7WldU3beF1bhqNqK94ijMaIj1dPiuRHqqfDt1o7ZimRIAANK47+svrHyNifNyqcWx8d/WX1j5GxPm5VOAJAdAP2QuF8m4r92Ef0gOgH7IXC+TcV+7ALGAAAAAAVsdOn2SeffB8J/D23DncenT7JPPvg+E/h7bhwAAAAOm9FrJ8rz/AI+aWyjOsvw+YZfib92m9hsRbiu3ciLFyYiYnlPOIn9Cwv6inCP8XGmf1fb+ZALod+yT0d8Ivfw91Z6Dn/1FOEf4uNM/q+38x9RThH+LjTP6vt/M6AA5/wDUU4R/i40z+r7fzH1FOEf4uNM/q+38zoADn/1FOEf4uNM/q+38xPBThFEbzw50xEeT7fzMhxS4laP4a5JOaaqzW3hoqj6zhqPT378+Kijtn3+yPDKAfH7pGav4oXb+WYWasj01M9WnAWK9671O8871f3Uz/ZjamNo5T2yHQekBxL4GZFRich4acOtI5vmfqKs1qy63XhbM895txt9dqjltPqfyo5TFOqZqqmqdt5nflG0P4AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA/XCYe/i8VawuFs3L9+7XFFu3bpmqquqZ2iIiO2U7eiv0Y8LpaMFrPiBhrGMz3qxewmXV0xXawUz2VVxPKq7HbHbFM845xEx+PQq4CUaewGF4j6vwlFebYuzFzK8Lcp3+hLVUbxdqj75VE8o+5ifHPKVgAAAPJnOZYHJ8oxmb5niKMNgcFYrxGJvV77W7dFM1VVTt4oiZBxPpj8YaeGuhYynJ8XFGp86oqt4WKKpivDWdpiu/vHZMTtTTzid53jfqztW/VM1VTVVMzMzvMz4W58a9eY3iTxIzXVWL61FvEXOphbNW31mxTyop5eHbnPuzLSwAAAAf2JmJ3idpTq6H/SOjUsYbQOvsbV9O6aIoy7M71W8Y2I5d7uVTz77t2VT6vad5irbrQUfdi7dsXqL1m5Xbu26oqoronaaZjsmJ8EguVEbeh90gLOvcusaM1ZjLdvVWFtbWL1yYp+mNumO2PHdiI3mI7Yiao7J2kkAppufZKvflcsppufZKvfkHys86HnsdNKfmLnnKlYazzoeex00p+YuecqB1wAAAAAFRnFT1z9VeWcX56tgcBiruCxtnGWO9d9s1xXR3y1Tcp3id43pqiaao9yYmGe4qeufqryzi/PVtaBveYcZOK+OxM4i5xF1RZmYiOphcyu4e3TEeCLduaaaY96IeKjifxKt3pvUcQ9XU3au2uM6xEVT+nrtRAdN01x94xZBXM4TX+dYmmqqKqqcwvfRkTt4N73WmI96Ydy4ZdNPMbOItYTiFp61isPVVtXjcsjqXKI585tVTtVty7Ko8M8+xEABb7orVum9aZFZzvS+b4bM8BdjeLlqedPuVUztVTPuTESzaqDgtxP1Hwr1fZz7IrvfLMzFGNwNyqYtYu14aavFPhirtifHG8TZ9w41hk+vNGZdqnI70XMJjbUVdXrRNVqvsqt1bdlVM7xINhAAAARb6dHBmxqPTl7iPp/B0051ldnfMqLVvni8NTHq6tu2q3Hh7erG08qY2lI/lyii5bqt3Kaa6KomKqao3iYntiYBTQOndJ3h9HDji/muS4axVayvEVfRmXelmKYs3JmYpiZ7Ypnenw+p8bmIAADL6O0znmr9R4TT2nMuvZhmWLr6lqzbj46qp7KaY7ZqnaIjnLyZJleYZ1m+EyjKsJdxeOxl2mzYsWqZqqrrqnaIiIWX9GfgrlHCXSlM3KKMVqXHURVmONmImaf8Awbfiop+Oqd5nwRAeHo49HzTvCnC05pi5tZtqi7a6t3G1U70WN/VU2YnsjwTVPOY8Ubw7WAAAEzERMzO0R2yrV6YfFeeJPEu9hMsxE3NPZJXXhsDtVvTeridrl6NpmJiqY9LPhpins32iYHTL4hf0D4NYy3g8TFrNs7qnL8HEVR14iqJm7ciO3amjeN9uU1U9m8K0QAAAAHaOh3xHu8PuL+CtYnFVW8lzuacDj7c1T1ImZ+tXZiPuqap5Tt2VVxy3lxd/YmYmJiZiY7JgFy8c43gcx6L2u6uIXBbI86xNya8xw9v6BzCZ7ar9qIpmvsiPT09WvlyjrbeB04BoPG3hVpnitpWvKM8w9NvF2omrA5hbp+vYW544nw0z4aZ5T78RMb8AqO4n6Hz7h3rLG6X1Dh+94rDVekuU/Y79ufU3KJ8NNUfpjnE7TEw1hZn0r+D+H4paBuXMvsWadS5ZRVey69MbVXYiN6rEz4qvBvyirb3VZ9+zdw9+5Yv2q7V23VNFy3XTNNVNUTtMTE9kxPgB8AAt90B9omn/ACZhvNUs2wmgPtE0/wCTMN5qlmwAAAAYzVv2qZv8BvebqU9LhdW/apm/wG95upT1HbzBmtE6Vz7WepMLp7TeXXcfmGJnai3RHKmPDVVPZTTHhmViXRt6P2QcKsvt5lj6cPm2q7lP13HzR6XD7xtNuzvziNpmJq5TV7kcnP8Aoma06PmiNL2sry7VuFwuocdT18xxmbWKsLVcqj7nvlX1umiPuaevz5z2ylJl+OwWY4W3isvxmHxeHuUxXbu2LsV0V0z2TExO0xPjB6AAAARY7o9qH6B4b6f05bv10Xc1zGq9XRFPpblqxRHW3nwbV3LUoHJK90P1DXmfGfA5FReuzYybK7dNVqr1NN67VNyqqn36O8xP5KNQJI9zu9fTG+Q7/nbKwdXx3O719Mb5Dv8AnbKwcAAAABDXulf2LRv5WJ/6EykNe6V/YtG/lYn/AKAQxd06NHR5zzipiKc6zSq5lWlbNzavEzH1zFzE86LMe54a55R4N53255wesaBva3wc8SMdjcLkFuqK71OFs1VzemJ9RVNM9ammfDNO87dm081lPDfiXwpz/L8JleitU5BNq1R3rDZfauU4e5RRRHZTYrimuKYjxU7A2fSGmsi0jp/DZDpzLMPluXYamKbdmzRtE+Oqqe2qqe2ap3mZ5zLLkTExExMTE9kwAAAKq+knqCdS8c9W5n3y/XbjMbmHtRe9VTTanvcU9vZHVnb3Fn2tM6s6c0dnWocTRXcsZXl9/G3KKPVVU2rdVcxHu7UqgcTevYnEXMRiLtd29drmu5crq3qrqmd5mZ8MzIPzABMrueHEquKsw4Y5piN7fpsdlPXqn0s/621HuTyriOXPr9u6ZyoHQ+pMy0fq/K9T5RXFONy3E0X7cTMxTXtPOirbn1ao3pn3JlbHoPU2Xay0dlWqMpqmcHmWGpv0RM86JmPTUT7tM70z7sAzYAAACEvdCeF1OCzDCcT8osxFrF1U4TNqKaZ9Ld2+t3uUbemiJpq3mOcUdvWlNpgeIOlst1rovNdLZtRvhMxw9VmqqO23VMelrjnHOmdp/QCoUZvXmmM00ZrDNNL5zbijHZdiKrNzq+priPU10/7tUTFUe5MMIAAAAA/bA4rEYHG2MbhLtVrEWLlN21cp7aaqZ3ifjh+IC1Do5cTMJxS4Y4HPqaqKczs/6NmliN97V+mOc9kelqjauNt49NtvvE7dIVm9EfirHDHibZnMr9dOn822wuYREbxa3n0l7baZnqz2xHPaZ7eULMaZiqmKqZiYmN4mOyQf0HL+k3xNs8LuFuNze1VRVm+M3wmV2Zq5zeqifTzG+/VojeqdvDtHLfcES+njxS/pbr6jRWVXoqyfT9cxerp7L+LmNq57OyiJ6keOetPONka36Ym9exOIuYjEXa7t67XNdyuud6qqpneZmfDMy/MAABI/ueHr84ryFiPO2UcEj+54evzivIWI87ZBYSAAAAACu7ugvsgP8Iw37a0eEh+6C+yA/wAIw37a0eAAAGS0znucaazzC53kOY4jLsxwlyLlm/Zq2qpmP+UxPZMTvExvExMMaAsR6K3SMwPEm1GmdVV2Mv1Xap3tbR1bWYUeGaPBFceGjwxO9O+1XVkOpuwOLxWAxtnG4LEXcNibFcXLV61XNNdFUTvExMc4mFkPRH40U8U9HTgM5u0f0pyq3TTjtqOrGJo7Kb8RHLedvTRG0RPZERMQDuAAAAPzxNixisPcw2Js271i7RNFy3cpiqmumY2mJieUxMeBX30v+j9e4f4+vWGksHVc0pirn1+1RznLrkzypmO3vdXgq57TG07b072EPFnuU5dnuTYvJ83wlvGYDGWqrN+xcj0tdExtMf8AcFOg6R0ieF2P4U8RcVkV3r3ssv74jLMVNMxF2xM8qZns69Pqav0T2TDm4AAOwdDH2TGj/wA5if4W8s4Vj9DH2TGj/wA5if4W8s4AAAAAQk7pb9sWivgmL/ftptoSd0t+2LRXwTF/v2wRCAAB6crwGMzTMsNluXYa5icZirtNmxZtxvVcrqnammPdmZB79HaazvV+o8Jp/T2AuY3MMXX1bdqiOzx1TPgpjtmZ7FjPRs4CZFwmyz6OxM2cz1RibcRiMdNHKzE9tu1vzinxz21beDsfXRd4I4DhNpib+Oixi9T4+iJx2KpjeLUfebc/2Y8M/dT70OygAAA07jPrvA8N+G+basxsUXK8La6uFsVVxT3+/Vyt249+ec7bzFMVTtOwI2dPjjHOHtTwr09f+uXaKbmd3qd46tM7TRYifDM+qq8G00x27xEK3tz3NcwzzOcZnGa4mvFY7G3qr+IvVbRNddU7zO0co96OUPEAAAAD0ZbjcZluYYfMMvxV7CYzDXabti/Zrmiu1XTO9NVNUc4mJiJiYWM9E7jthOKOQ/SXPL9qxq3A2479b5UxjLcf66iOzf8AtRHZPPslW8yGnM7zXTmeYTO8jx17AZjg7kXbF+1O1VFUf8pjxxPKY5SC4gcY6MXHXK+LeRTg8ZFjAaqwVqKsbg6J2pu08o79aiefU323jnNMzETM7xM9nBpXHf1l9Y+RsT5uVTi2Pjv6y+sfI2J83KpwBIDoB+yFwvk3Ffuwj+kB0A/ZC4Xybiv3YBYwAAAAACtjp0+yTz74PhP4e24c7j06fZJ598Hwn8PbcOAAAAB1vod+yT0d8Ivfw91Z6rC6Hfsk9HfCL38PdWegA1jiTr7SnDzTt3PNV5rZwWHpie9Wt+texFX9i3R21VdnZyjtmYjeQbNXVTRRNddUU00xvMzO0RCM/SG6VeRaOuYjT2hYsZ7ntMTRdxe++FwlcTtNO8fZKo2nlHKOXOZ3iI+9ITpMaq4jVXckyKbmQaa3qpmzar+v4uN+U3a47I5eop2jnO81ctuBAzOsdUag1hn1/PNS5tiszzC/Pprt+uZ6sbzMU0x2U0xvO1MbRHghhgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAASJ6EnB+Ne6zq1XnmFi5pzJLsekrpiqjFYraJptzE9sUxMVVcv7MeFwPI8sx2d51gsnyyxViMbjsRRh8Pap7a7ldUU0x+mZha/wh0VguHvDvKNJ4KaaowVmIvXI32uXavTXK+fjqmZBtgAAACNHdBdc16f4X4TSeEuTTidQ35pu7VbTGHtbVV+Dw1TRHbH6Ul1anTb1XVqjpAZvZor62FyW3byux6XqzHe96rm/j+u13OfhiIBxIAAAAAAAH74DF4rAY6xjsFiLuGxWHuU3bN23VNNVuumd4qiY7JiVhXRJ6QNjiRltvS2p67eH1bhLfK5vEUZjbj/WUx9zciPVU+HbrRynq013PZkuZ5hk2bYXNsqxd7B47CXabti/aq6tduuJ3iYkFximm59kq9+VlHRZ46ZZxU07TluZ3rOF1bgrf+l4afS/RNMf66344nwxHOmfcmJmte59kq9+QfKzzoeex00p+YuecqVhrPOh57HTSn5i55yoHXAAAAAAVGcVPXP1V5Zxfnq2tNl4qeufqryzi/PVtaAAAAASv7nfxAxOX6xzDh7jb1VWAzS1Vi8FTVMz3rEW49NEeCIro3mfdop8cooNy4H5vdyLjBpLNLPe+tazbD0zNfqYpqriiqZ/RVILZgAAAAARO7o7o+jG6MyHW+Ht1fRGWYucDiepb33s3omqmqqrwRTXRtEeO9P6YMLQel1ldjNejvq21iOvth8LTiqOr/bt101U/o3hV8ADpnRn4dTxM4tZbkV+3NWWWP9MzKd5j6xRMb07x2daZpp8Hqu2ASg6BvB61kmQUcTM+wn/1TMrUxlVNz/UYartuRExyqr8E/wBns5VTvKx82bduzZos2bdNu3RTFNFFMbRTEcoiI8EPoAAACZiImZmIiO2ZBXf0+tYV6h40RkNq71sHp/DRh6aYmdovV7V3J2nsn1FPLwUQjuzevc/u6p1tnWo73fetmWOu4mKblXWqopqrmaaZn3I2j9DCAAAAAAAl93NzVNdvPdS6Mu3Jmi/h6MxsU9XsmiqKK+fg9XRy99NpVv0VNQ16a6QWkMfG827+OjA3aZr6tM0X6Zs7z44pmuKtp8NMLSAAAFeXTv4bf0Q4nUaoy3DRRk+o6arvpKdqbWKp277T7nWiaa47N96oiPSysNcc6ZGj6dX8B866lnvmMymmMyw+0RvHe4nr9vZ6SawVlAAt90B9omn/ACZhvNUs2wmgPtE0/wCTMN5qlmwAAAAYzVv2qZv8BvebqU9LhdW/apm/wG95upT0AzOltU6k0tjIxenM8zDKr0VRX1sLfqoiqY7N4idp/SwwCUPCXpi6vyO9bwWvcDb1Hl3Z9E2YptYy3G0R+RcjlPKYiZmfVeBMThdxQ0TxJy+cVpXOrGKu0U9a9hK56mIsxy3mq3PPbnEbxy59qpp79PZ1m2ns4w2cZHmGJy/MMNXFdnEWK5pron348HjjsmOUguJETOjj0scFnVdjTXE25Zy/MJpijD5vEdWziKt9urdiOVurs9N6mdp32nbeTOts8w+ntD53qS7TVew+W5bfxtUWpjeum3bqr2pns3mI5ArC6R2e/wBI+OGrcziu7VbnMrlm33yedNNuepEe9HV5OfP0xN+9icRdxOIu13b12ua7ldc71VVTO8zM+GZl+YJI9zu9fTG+Q7/nbKwdXx3O719Mb5Dv+dsrBwAAAAENe6V/YtG/lYn/AKEykNe6V/YtG/lYn/oBDEAHRNA8bOJ+iblH0k1bj5sUcvobFV/RFmY22iOrXvtER2bbbJW8H+mLprPIs5dxBy+NPY7bacbh+tcwlyfHMc67fvemj3fAgcAuQy3HYLMsDZx+XYzD4zCXqetav2LkXLdceOmqOUx7z0KseCnGrWvCrMI+kuOqxOU13IrxGV4iqarFzt3mn+xVz7adt+W++0LB+B3GPSXFjI6cVkuI+hs0tWoqx2V3qo79h57J2/t0b9lcdsTG8UzygNZ6b2eTkvR2zy3R36LmZ3bGBpqt1dXq9auKqt/cmmiqmY8PWVqJr90pzvveUaR07TTXvfv38ZVVFfLaiKaIiafDzr339xCgAABMzud/EqnfMOGOaX5iqetjspmrw7fZrXv9lcR+X4oQzZrQupMfpDWGV6myyqacVl2Jov0Rvt1tp50z7kxvH6QW/DD6J1FlurtJZXqbKL0XsFmWGov2qo7Y3jnTPiqpnemY8ExMMwAAAACH/dDuGdWJy7AcTcqw+9eF6uDzaKe3vcz9au7beCqZomd/uqNo7ZQnXEalybAai09j8izSzF7BY/D14e/RPhoqjaVUXFvROZcPOIeb6SzOirr4K/MWbsxyv2Z527kbTPKqmYnbflO8TziYBqgAAAAACwzoM8VqdaaA/ohm2Iqqz7ILcURNc7zfwm+1uuOXbT6iY59lM7+m2ivNLDuc+ir2P1jnOu8RbuUYTK7H0Fhq9pim5fuxvXET2T1aIjeP/EpBOhWn0w+KX1SOKN/D5bfrqyDJZqwmBjfldqifrl7b/eqjaP8AdiOzeYS26aXFGjh/wvuZVl9+mM+z+K8LhqY7bVnb67dn3omKY8c1eHaVboAAAACR/c8PX5xXkLEedso4JH9zw9fnFeQsR52yCwkAAAAAFd3dBfZAf4Rhv21o8JD90F9kB/hGG/bWjwAAAAA2vhJrjNOHev8ALNV5Tcqi5hLu161E+lv2Z5V26o8MTHxTtPbES1QBcPpjOsBqLTuX59ld3vuCzDD0YixX46KoiY/ayKKHc69eXMy0hmugMde61zKLn0VgImOcWLtUzXR2eC5M1eP65PghK8AAAAHGel5wvo4kcKsXVgrNM59k9NWMy+rlE3OrG9yzM7T6qmJ2jl6aKd5iN1Zs8p2lcwrH6X2gZ0DxrzOxh7M0ZZm3/wBSwM77x1bkz16ffpuRXG3bt1Z8IOPgA7B0MfZMaP8AzmJ/hbyzhWP0MfZMaP8AzmJ/hbyzgAAAABCTulv2xaK+CYv9+2m2hJ3S37YtFfBMX+/bBEIABNjoC8HqcNgp4p6gw+9/ERNrJbNceot84rvzEx21epp8URVPPrRtF/gVoHFcSuKGUaVsTNGHvXe+469G/wBaw1HO5V2Tz29LTvy61VO+0c1q+VYDCZXlmGy3AWKbGEwtqmzZtU9lFFMbRHxQD0gAAAIOd0V15GO1RlPD7AYmarWW2vozMaKZ5d+uR9bon3Yo9N4vrlKcGIvW8Ph7mIvVxRatUTXXVPZERG8yqR4r6kvav4lah1Lfu98nH4+7conrzVEUb7UREzz6sUxTER4IiIBrAAAAAAAAMnpbPs20xqHBZ/keMuYPMcFdi7YvUTtNM+L3YmN4mPDEzCzHo4cY8n4t6RpxFuqjC59g6KacywO/Omrs75R46KvBPg7JVds/w/1hqDQmqsJqXTWOrwePw1XKY50XKJ9Vbrp7KqZ8MT7/ACmIkFonHf1l9Y+RsT5uVTixjCcYMg4t9GvV+PwNVvCZxhskxFOZZdNXp7Ffe59NT4ardXgq/RPOJVzgJAdAP2QuF8m4r92Ef0gOgH7IXC+TcV+7ALGAAAAAAVsdOn2SeffB8J/D23DncenT7JPPvg+E/h7bhwAAAAOt9Dv2SejvhF7+HurPVVvRp1DlGlON+nNRZ/jKcHlmArv3cRemmaurT3i5HKIiZmZmYiIjnMzEOjdIPpSak13Tish0lF/INOXOtbrqirbFYuiY2mK5iZimmefpaZ7J2mZBIPpB9KLTOg6MVkWk+85/qOmKrc1UV74XCV7dtdUermJn1FPimJmEENeaz1NrnPrud6ozbEZjjLkzMTcq9JbifuaKY5U09nKPE18AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABJvue2iKc84nY3V2KsRXhsgw+1mqZ5RiLsTTTPb2xRFfbExz8eyfrhXQZ0lGmeAWXY27broxmfX7mZXorpjeKJnqWoiY7aZt0U1xv4a5d1AAAAB+GYYinB4DEYuqmaqbFqq5MR2zFMTP/sp/wBT5nXnOpMzzeuq5VVjcXdv73KutV6aqZ5z4Z5rWeNeOxOV8HtY5jg7ne8Thcjxl21Xt6mqmzVMT8apUAAAAAAAAAAGQ05nebaczzCZ3kePvYDMcHdi7Yv2p2qoqj/lMeOJ3iY5TvDwTO87z4X8AFnnQ89jppT8xc85UrDWedDz2OmlPzFzzlQOuAAAAAAqM4qeufqryzi/PVtabLxU9c/VXlnF+era0AAAAA+qKqqK6a6KppqpnemqJ2mJ8b5AXK2PsFv8mP2Pt8WPsFv8mP2PsAAAAGo8acDRmXCPVmCuRM015RiZmPetzV/7KlVu/E31t9T+R8X5mtUQAsI6AOhKNOcJ72qsTY6uY6jvd8iuZ5xhre9NuI57c5murflM9aN+yEAsmy/F5vnGCynAWK7+LxuIt4exao9VXcrqimmmPdmZiFvOkskwum9L5Xp/Bf8A2+XYS3hrc9WKetFFMR1piOW87bz7sgygAAADTuOOOnLeDGtcbRiJw9y1kONm1cidppuTYrijb3etMbNxcr6XFyq30ctZ1UTtM4Oin9E3aIn/AJSCrsAAAAAAG16M4b691lFFemdI5xmVm5M004i1hqosbx2x32dqIn3JkGv5NivoHOMFjZmqIw+Iouz1e30tUT/7LhMpxlGY5VhMwtxMUYmxRepifBFVMTH7Vf8ApHodcT82pou5xicoyK1VTTVtevTduRv2x1aImN49/wDSn1p3AXMryDL8su34xFzCYa3YquxR1YrmmmKd9t52327Nwe8AB+eKw9jFYW7hcTaovWL1E27luunemumY2mJjwxMP0AVAa5yK5pjWmd6buXJu1ZXmF/Bzcmnq9873cmnrbeDfbf8ASwzs3TUym5lXSL1HVX1Orje84uiKY22iq3THx70y4yC33QH2iaf8mYbzVLNsJoD7RNP+TMN5qlmwAAAAYzVv2qZv8BvebqU9LhdW/apm/wABvebqU9AAAAAOrZVxz1dZ4M51wwzPE3cwy3G2LdrBX67kxewdNN2iqq3FXbVbmmmaerPZE7Ry5OUgAAJI9zu9fTG+Q7/nbKwdXx3O719Mb5Dv+dsrBwAAAAENe6V/YtG/lYn/AKEykNe6V/YtG/lYn/oBDEAAABldJaizrSmoMJn+nsxv5fmOEr69q9aq2mPHE+OmY5TE8pidpYoB0PjvxSzHizqLKs+zXBWsHi8HlVvA3qbVUzRcrpuXK5uRE+p375HL3HPAAAAABM/udvEaKreZcM8xxHOnrY/LIqntjl323H/Kvb8qfGmSqD0DqbMdG6zynVGVXJt4zLcTTfo27KojlVTPjiqmZpmPDEytl0XqDAar0plmpMrr6+DzHDUYi1O/ZFUdk+7E7x+gGXAAAARY7oHwzrz3SGG4hZVhZuY3JKe9Y+LdETVXhaquVU7RvMUVTv7kVVTyiJlKd+GYYTDY/AYjA4yzRew2JtVWr1uuN6a6KomKon3JiZBTcN8496AxXDXijm2mL1uuMLRX3/AXKqZiLuGrmZoqiZmd9udMzvPOmqO2JaGAAAAD+00zVVFNMTMzO0RHhWldHXSWG4a8Dcny7H1W8HXbwtWPzK7e2txRcrjr1zXMxG3VjaPTdkU7T2IH9ErQVOv+NeT4PF2K7uVZdX9MMftTvTNFud6KKt4mNqq+rTMeGmatknun3xQjTei7GgcrxE05rntvvmL6vKbWDiqY7f8Afqpmnl4Kav0hE3pGcSb3FHihmGoKartOW25+hsttV7xNGHpmerMxvymrnVMR4Zc4AAAAABI/ueHr84ryFiPO2UcEj+54evzivIWI87ZBYSAAAAACu7ugvsgP8Iw37a0eEh+6C+yA/wAIw37a0eAAAAAAAdc6IGqZ0r0gNN3aqpjD5nf+ll6Ip3me/wDpKPe+udTefFus8U24LEXsHjLGLw9yu1es3KbluuidqqaqZ3iYnwTEwuD03mdrO9PZbnNiiq3ax+EtYqiirtppuURVET8YPeAAAAjJ3QzSFOccLMBqmzZmvE5Hi9q6qaZmYs3tqat/BEdaKOfvJNtX4t6ZjWXDHUml4t2q7uY5des4eLkzFNN7qzNqqdvFXFM/oBUeADsHQx9kxo/85if4W8s4Vj9DH2TGj/zmJ/hbyzgAAAABCTulv2xaK+CYv9+2m2hJ3S37YtFfBMX+/bBEIH1boquV00UUzVVVMRTERzmfECcXc59DzgtKZxr7GWK6buZXvoLBTXTMRVZt8666Zntia5mnePDbqhLRqXBrTdnSPCrTWnLNNFP0Fl1qm5NFHViq5Mda5Vt4JqrqqmfdmW2gAAAA0TpCZ/8A0Z4JavzqMV9C3bOV3aLF3bfq3rkd7t8vdrrpj9KqJY909M3tZb0dMywdyJmrNcdhMJb9yqLkX/2WZVwgAAAAAAAAAAyOQ55muRX8ReyrGXMNVicNcwt/qzyuWrlM01UVR4YmJY4AEgOgH7IXC+TcV+7CP6QHQD9kLhfJuK/dgFjAAAAAAK2OnT7JPPvg+E/h7bhzuPTp9knn3wfCfw9tw4AAAAAbXlnDnW+ZaIxutcHprMLmn8HT17uOm31bc0xVNNVVG/q4pmJiqad4p2nfbaWqAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPq1bru3aLVumaq66oppiPDM9j5ZXR1E3NXZPRTG81Y+xEf/yUgtm4eZXYyTQeQ5RhrVVqzg8usWaKKp50xTbiNmdAAAAAGi9IWZjgTrrbt+kGM8zUqhWycd7N3E8E9b4exbqu3bmQY2miimN5qmbFe0QqbAAAAAAAAAAAAAWedDz2OmlPzFzzlSsNZ50PPY6aU/MXPOVA64AAAAACozip65+qvLOL89W1psvFT1z9VeWcX56trQAAAAAALlbH2C3+TH7H2+LH2C3+TH7H2AAAADXuJvrb6n8j4vzNaohbvxN9bfU/kfF+ZrVEA6v0Rsmt530idI2L9iu9Zw2LnG1dX7ibNFVyiqfciumhaCrt7n3Tv0gImY32yjEz2dnOhYkAAAAA5t0ocFczDo+a1sW6etVTldy9Me5bmLk/8qZdJaxxay7FZxwq1dlOBt98xWNyPG4exR/arrsV00x8cwCo4dJ0Zwtweb9S5n3EnQ+nLM1zTXRfzSjEXojblVFNrrUTE+7XE+47FpDg70bcHbor1Pxpwea3epHWow2Kt4e3FW/OY9VO3uf8wRUbRpTh5rnVXVnT2k84zGiuma6blrC1dSqInaZiqY2n407NE4nom6NrpvZDj9EWsTRci5RicRfjE36Kojbem5dmqqj3qZiG/W+N/B23RFFviFpyimOyKcVTEQCHOjuh5xUzi5buZzdyfT2Gmqnr/RGIm9e6s9s00W4mJmPFVVS7Vo/oW6Dy6KK9SahzfPbsRMV02qacJaq8UxTE1VRt+XLr31cuEH4xdO/K6T6uXCD8YunfldIP10hwX4W6Umi5k+icopv0dWYv4ixF+5FVPZVFVe/Vn3Y2b+539XLhB+MXTvyuk+rlwg/GLp35XSDog539XLhB+MXTvyuk+rlwg/GLp35XSDog539XLhB+MXTvyuk+rlwg/GLp35XSDog539XLhB+MXTvyuk+rlwg/GLp35XSCHndDctnCcbsLj5mdsdlNqqI8XUqqp/8AZG5Izp56v0xrHiDkGN0tnmBzjDWMqm1du4W7FdNFffa56szHh2mJRzBb7oD7RNP+TMN5qlm2E0B9omn/ACZhvNUs2AAAADGat+1TN/gN7zdSnpcLq37VM3+A3vN1KegAAAAAAAASR7nd6+mN8h3/ADtlYOr47nd6+mN8h3/O2Vg4AAAACGvdK/sWjfysT/0JlIa90r+xaN/KxP8A0AhiAAAAAAAAAAAAmj3O7iRXcw+Y8MszxHWi11sdlXWn1NMz9dtx7m8xXEeOa/Ghcz3D3VWZaJ1rlOqsoqiMZluJpvU01epuR2VUVe5VTM0z7kgt6GI0XqHLtWaUyzUmU3Ovgsxw1F+1vMTNMVRzpnblvE7xPuxLLgAAAAjp07OGVOr+Gs6ty7DU1Zxpyiq9VVTT6a5hO25TM7c+r6uPF6bxzvXkuXqpiqmaaoiaZjaYmOUquOk7w2q4Y8VswyfDWaqMoxU/ReWTzmIs1TPpN55z1J3p7Z7I5g5eAADL6LyHGao1blWncBRVXicxxdvDW4jbfeqqI357RyjnzBNHoPZBl/D/AIJ57xT1LMYK3jouXabtydtsHY3iJiO3eqvrxEds7U7b9aEQOLGtMfxB4g5tq3MY6tzHXt7dv71aiOrRR+imIhJXpxa2wemtLZFwP0ze6mGwOEw9WY027k+lt0Ux3mzVz38EXJir/cnxSiCAAAAAAAkf3PD1+cV5CxHnbKOCR/c8PX5xXkLEedsgsJAAAAABXd3QX2QH+EYb9taPCQ/dBfZAf4Rhv21o8AAAAAAALV+jdmN/NeBGjcbiZ3u1ZXaon3qPSR/yphVQtL6KvseNFeTY/eqB00AAAAAFSHFzK4yXinqnK6MNGGt4bN8TRatRHKi332qaIj3OrMNWdR6WERHSK1pEREf6fHm6HLgdg6GPsmNH/nMT/C3lnCsfoY+yY0f+cxP8LeWcAAAAAISd0t+2LRXwTF/v2020JO6W/bFor4Ji/wB+2CITbeDWU0Z7xZ0pk925Vaoxeb4a1VXTG80xNynm1J17oa4PD47pLaPsYmjr0U3sReiP9+3hrtymf0VUxILOgAAAAARq7ox6xuV/3isfw+IV+LGunzk8Zn0d8bjZrmn6U5jhcZER911q5sbf/wC/f9CuUAAAAAAAAAAAABIDoB+yFwvk3Ffuwj+kB0A/ZC4Xybiv3YBYwAAAAACtjp0+yTz74PhP4e24c7j06fZJ598Hwn8PbcOAGZ0fpXUer84tZRpnJsZmuNuTERaw9uatt/DVPZTT46pmIjwymFwZ6G+X4WizmfE/HRjr0xFU5VgrtVNqnt9LXdjaqrwep257857ZCK3DLhjrXiPmUYPSmSX8XTFXVu4quOph7Pu1XJ5R29nOfcTT4K9ErR+k7eHzPWly3qfOIpprqs1UTGDs17RM000zzuRE7+mq23jb0sdiQ+S5VleSZbZyzJ8uwmXYKxHVtYfC2abVuiPcppiIh7Aee9gcFey6vLb2EsXMFcszYrw9VuJt1W5jqzRNPZ1duW3ZsrZ6WfBzEcLNczicuszVpjNq6rmX3KYnaxV21WKpnw09sc+dO3hiYiy5rHFLRGT8RNEZhpTO6aow2Lo9Ldo269m5HOmunfwxIKjhsvEzRWecPtZ4/S2oMP3rF4S5MU3I9Rftz6i7RPhpqjn447J2mJhrQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADN6Dqpo1xkVdXZGY2Jn/wDkpYR7tP4i3hM/y/F3pmLVnFWrlcx4qa4mf2AuKH44HE2sZgrGLszM2r9um5RM/wBmqN4/a/YAAAAHlzjDVYzKcZg6dutfsV2437N6qZj/AN1PebYO5l2aYvL7tVNVzC367Ncx2TNNUxO3xLjlXXSv0xXpPj5qfARZi3hsTiYx2G6trqUTbvRFe1MeGKapqo3jw0yDlgAAAAAAAAAAACzzoeex00p+YuecqVhrPOh57HTSn5i55yoHXAAAAAAVGcVPXP1V5Zxfnq2tNl4qeufqryzi/PVtaAAAAAABcrY+wW/yY/Y+3xY+wW/yY/Y+wAAAAa9xN9bfU/kfF+ZrVELd+Jvrb6n8j4vzNaogEi+573KaePFyid96soxG36KqFhqt/oH5nYy/pEZbZvRPWx+CxOFtbR911Ov+y3KyAAAAAAAFP2tcnnT2sc6yHr1V/S7H38LFdUbTVFuuaYq292I3/SxDuPTh01Xp7j/muIiiuLGb2bWPtVVbbVdaOrVtt4Iqoqhw4AAAAAAAAAAAAAAFvugPtE0/5Mw3mqWbYTQH2iaf8mYbzVLNgAAAAxmrftUzf4De83Up6XC6t+1TN/gN7zdSnoAAAAAAAAEke53evpjfId/ztlYOr47nd6+mN8h3/O2Vg4AAAACGvdK/sWjfysT/ANCZSGvdK/sWjfysT/0AhiAAAAAAAAAAAAACanc8OJVN3CZhwyzTEfXbXWxuUzXVHpqP9baj3YnauI58pr7IpTEVDcPNVZjonWuVapyuqfonL8RTdiiKurFymJ9NRM8+VUbx2T2rY9H6gyzVelst1Jk1+L+AzHD0YizXHbtVHZPiqid4mPBMTAMsAAAA4h0zOGdriBwmxONweHpqzvIYqxuDrjaKqrcR9dtbz4KqY325emop92J7eApnHa+mHwvjhxxTxF7LcPVbyHOpqxeB2j0tqqZ+uWon/dqneI/szEc9t3FAEg+h/g8t0xRqrjRqCx3zL9J4KbeBiZmOvjb3pKYjae3q1dXnG312J7ad4j46pxbzSrT+hdOcJ8NNNFWWx9M89ppnnOY3ad+91e7Zt1RbmPBV1wc/1bn2Y6o1PmWoc3v1X8dmGIrxF6urw1VTvty5REdkRHZEQxYAAAAAAAJH9zw9fnFeQsR52yjgkf3PD1+cV5CxHnbILCQAAAAAV3d0F9kB/hGG/bWjwkP3QX2QH+EYb9taPAAAAAAAC1Toz4K/l/ATRmFxNPVu05XbqmPcq3qj/lMKrrNuq7eotURNVVdUUxERvMzK4LSeV/SPS2U5L33vv0vwNnC9829V3uiKd/07AyYAAAAAKuOlj7IrWnw+PN0OXN0465ndzjjNrDML1VNVVecYmiJpneJpouTRT/yphpYOwdDH2TGj/wA5if4W8s4Vj9DH2TGj/wA5if4W8s4AAAAAQk7pb9sWivgmL/ftptoSd0t+2LRXwTF/v2wRCdm6EvsndI//AL3+CvuMumdFjMr2U9IbRWKserrzKnDT+TepqtVf/wBa5BaYAAAAADnPSbyOjUXALWWWVW712r6W14m3RaiZqquWZi9RERHb6a3TyVWrkcww1vG4DEYO9v3u/aqtV7du1UTE/tVD64yS9pvWWc6fxFE0XMux13DVUzVEzHUrmO2O3sBhgAAAAAAAAAAAEgOgH7IXC+TcV+7CP6QHQD9kLhfJuK/dgFjAAAAAAK2OnT7JPPvg+E/h7blWg6dN16yyqjV84uMhqxNEY6rC19W5Tb35zE7T2eHaN9t9tp5uq9On2SeffB8J/D23DgW6cPNKaP0pp3DYTRmU4HAZbctUV26sNTvN6nqx1a6q59NXMxtPWqmZntbIiB0BuMNWNw0cLNQYua8RYoquZJcuVTNVduImquxvP9mImqmPBG8dkRtL8AAAAHD+lxwYscUNFzmGU2LVvVOVUzcwl3qzvibcRM1Yerbt37aZ2naY25RVKtq9buWb1dm7RVbuUVTTXTVG00zHKYlcqhX07uCn0PiL3FLS+Crm3dq3zyzapiaaKuURiIiOcb9lXu+m5bzMhDsAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFsPAXUEao4NaTzzv1u9XiMstRdqtxtTF2iOpcp/RXTVH6G7or9zo1hOY8Pc30ZisTNd/J8Z9EYWiqv1OHvc5ppjxRciuqfduJUAAAAAIf90d0TOIy3INf4WzM3MLM5bja4+91TNdrfn4KpuRyj7vn2QmA1zibpHL9eaDzjSWZz1bGY4aq1F2ImZs3O2i5ERMbzTVFNW2+07bTymQVFDKasyHMtL6lzDT2cWJsY/L8RVYvUTExG9M7bxvEbxPbE+GJhiwAAAAAAAAAAFnnQ89jppT8xc85UrDWedDz2OmlPzFzzlQOuAAAAAAqM4qeufqryzi/PVtabLxU9c/VXlnF+era0AAAAAAC5Wx9gt/kx+x9vix9gt/kx+x9gAAAA17ib62+p/I+L8zWqIW78TfW31P5Hxfma1RANw4Kamp0dxb0tqS5fqsYfBZnZqxVymjrTGHqq6t7aPD9bqrhbQpnjlO8LUujXq+nW3BXTmdTX18RRhYwuJ3mZnvtr0lUzM85mdt9/dB0YAAAAAEX+6GaE+nfDrLtaYO1E4zIL828R1aedeGu7RO8xG89WuKZjedoiquUB1xGpMnwGodP4/I80sxewWPw9eHv0cudNUbTtv4fDE+NVJxd0LmnDniBmelM1pqmrC3d7F7qzFOIszzouU7+CY7fFMTHgBqQAAAAAAAAAAAAALfdAfaJp/yZhvNUs2wmgPtE0/5Mw3mqWbAAAABjNW/apm/wG95upT0uF1b9qmb/Ab3m6lPQAAAAAAAAJI9zu9fTG+Q7/nbKwdXx3O719Mb5Dv+dsrBwAAAAENe6V/YtG/lYn/oTKQ17pX9i0b+Vif+gEMQAAAAAAAAAAAAAE1+53cR4xGAzLhpmWJ+u4brY7K4rqneq3M/XbdO/LlMxVt/vVTtymUKGf4eaox+itb5RqrLK5pxWW4mm9TETMdensqonaY5VUzVTMeGJkFvIxekc9wOp9L5ZqHLLnfMHmOFt4mzVt9zXTE+Hw89mUAAAAByjpU8NLXEzhPjsBh7NM51l0TjcrudWOt32mOdreeymunent236szv1dlYF23Xau1WrlFVFdFU01U1RtMTHbErlldvTm4XxoniX/SbK7E05JqKar8RTRPVsYqPslG/ZtV6uPfqiI2pBwrTWNsZbneGzK/Zt34wlff6LVynem5XTzpiqNpiaetEbxPKY3jwvLmGMxWYY/EY/HYm9isXibtV6/fvVzXXdrqmZqqqqnnMzMzMzPbu/AAAAAAAAAASP7nh6/OK8hYjztlHBI/ueHr84ryFiPO2QWEgAAAAAru7oL7ID/CMN+2tHhIfugvsgP8ACMN+2tHgAAAAAAHUuilpivVfHzSuDm3VXh8HjKcwxExTExFFj65HWifBNVNNM/lLRkPO5yaGuWcvzziDjbFVEYir6XYCa6NutTTtVdrpmY7N+rTvE7b01x4EwwAAAAGD4gahs6S0Nnmp79FNyjK8BexcW6q+rFyqiiaqaN/BNUxFMe7LOI490A1hRkXB23py1eppxWf4qm1NG8bzZtzFdc7Tz236kbx448YK97lddy5VcuV1V11TNVVVU7zMz2zMvkAdg6GPsmNH/nMT/C3lnCsfoY+yY0f+cxP8LeWcAAAAAISd0t+2LRXwTF/v2020JO6W/bFor4Ji/wB+2CITJaXzKrJtS5Zm1FdyicHi7V/e3O1UdWqJ5e7yY0BchluLt4/LsNjrO/e8Rapu0b+KqImP2vQ470N9Y0aw4CZHXXP+mZRE5VioiJ9VaiOpO8zO+9qbczP9qavE7EAAAAAr+7oLoWch4q4fV+EtbYLUOHib0xHKnFWoiirsjaIqoi3V2zMz15WAuW9KLhvPE3hNmGUYO3TVnGE/0zLN+rHWvURP1veeyK43p7Y5zEzO0SCrkfd63cs3q7N63XbuUVTTXRXG1VMxymJieyXwAAAAAAAAAAAkB0A/ZC4Xybiv3YR/SA6AfshcL5NxX7sAsYAAAAABWx06fZJ598Hwn8PbcOdx6dPsk8++D4T+HtuHA9WUZhjMpzXCZpl2IuYbGYO9RfsXrdU01W66ZiqmqJjnExMQtG6OnFDAcVuHOFzy1XRRmmH2w+a4ammae834jnMRPbTVG1UTG/bt2xMRVe6b0buKeK4U8RsPnNXfLuUYrbD5pYojea7Mz6qmN49NTPOP0x4QWljzZVj8HmuV4XM8uxFGJweLs037F2j1NyiqImmqPfiYekAAB+GYYPC5hgMRgMdh7WJwmJtVWb9m7TFVFyiqJiqmqJ5TExMxMP3AVkdKjg/iuFGu6qcJbm5pzM5qvZZejeYtxv6axVM/dU8vHvTNM9u8Rx5bVxd0Dk/ErQuO0rnPWt28RT1rOIopia8Pdj1Nynft2ntjwxMxy7VWvEPR+d6E1fj9Magws2Mbg7nVmfubtH3NyifDTVHOJ+PnEwDXwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAdZ6J/EG3w74y5XmGOvzayjHz9A5hVO+1Fu5ypuTzjlTX1ZmefpettG6z9TOsi6F3FGxr/hfYyfG4mJz/ACCijC4q3VtFVyzEbWrseOJiOrM+CafdiZDuwAAAAAImdPDgxczvL44maaw1dzMMFa73m2GtW95vWI7L0bc+tR2Vdu9MxPLq+mg0uXuUUXKKqK6aaqKomKqao3iY8UoCdLbo54zR2YYjWWiMDdxOmsRXNeKwlqmaqsuqnn2ds2p8E/c9k+AEYwAAAAAAAB+mGsX8ViKMPhrNy/euT1aLdumaqqp8URHOZZLVGnc40xmMZZnuDrwOP73Tcrw1zlctRV2dePuZmOe089pifDAMSs86HnsdNKfmLnnKlYazzoeex00p+YuecqB1wAAAAAFRnFT1z9VeWcX56trTZeKnrn6q8s4vz1bWgAAAAAAXK2PsFv8AJj9j7fFj7Bb/ACY/Y+wAAAAa9xN9bfU/kfF+ZrVELd+Jvrb6n8j4vzNaogBK3uefEWnJ9V5jw9zG9FOEzn/SsBNVW0U4minaqmPy6Ijw9tuNo5yik9uRZrj8jznB5xleJrw2OwV6m/Yu0TzprpneJ/7AuLGicCOI2X8UOHGX6mwvUtYuqiLWYYan/UYimI69Mc59LM86Z8Ux4d29gAAAAOFdLvgpTxR0lRmeSUWrWqMqpmvDTNP/AN3a2nrWJnwT4aZ58425RVvHdQFNmLw9/CYq7hcVZrs37Nc27luuNqqKonaYmPHu/JYT0pOjZgOIFvF6r0fbsYHVUUzcu2Z2otZhMeCqeym5MdlXZM7dbbeaogNqLJM207nGIyfPMvxGX4/D1dW7Yv0TTVTPvT4PdBjwAAAAAAABsvD/AEHq7XubRlmlMjxWZX94iuqina3b38Ndc7U0x78tr42aCyPhdTg9JXMzozrV1dFN/Nrtif8ARcBTO/VsW/uqq57apqiNo6u0emnYOXgAt90B9omn/JmG81SzbCaA+0TT/kzDeapZsAAAAGM1b9qmb/Ab3m6lPS4XVv2qZv8AAb3m6lPQAAAAAAAAJI9zu9fTG+Q7/nbKwdXx3O719Mb5Dv8AnbKwcAAAABDXulf2LRv5WJ/6EykNe6V/YtG/lYn/AKAQxAAAAAAAAAAAAAAABNfudvEavEZdmPDPMb81ThetjsrirwW6p+u24nfsiqYriNu2qrnO/KYKobh1qrMtD64yjVmU1RGLy3E03qYmN4rp7K6J9yqmaqZ8O0ztstl0ln2X6o0xluosquTcwWYYejEWZnbeKao32naZjeOyefbAMoAAAA4n02cPpq90f85uaiqmi7aqoqyyqimJr+it/SRHOOUx1oq/3d55zEQ7Yru6cfFP+m/EedMZVfivItPVTZpqpq3pxGJn7Jc7I5U+ojt9TMxO1XII8AAAAAAAAAAJH9zw9fnFeQsR52yjgkf3PD1+cV5CxHnbILCQAAAAAV3d0F9kB/hGG/bWjwkP3QX2QH+EYb9taPAAAAADNaH01mesNXZZpjJrNV7HZjiKbNqIjfbfnVVPuU0xNUz4IiZYamJqqimmJmZnaIjwrAehJwUvaGyGrW2psJTa1Bmtnq4azV6vCYaradqvBFde0TMdsRtHbvEB3nh9pjL9GaKynS2V0RThctw1NmmYjbrT21VT7tVUzVPuyzoAAAAAKzumLxEq1/xkx0YW/wBfJ8licvwNMbxE9Wfrtzt7aq9+fL0tNET2JgdMbixTw44b3cvyy/RGoc7pqw2EjfnYtzG1y9tt4InaN/DVE89phWwAADsHQx9kxo/85if4W8s4Vj9DH2TGj/zmJ/hbyzgAAAABCTulv2xaK+CYv9+2m2hJ3S37YtFfBMX+/bBEIAEme5+a+tae4k4vR+PvxbwmobcRh5qriKYxNveaY7O2qmaojnHPaOczCf6m/LMbi8szLC5lgMRcw+Lwl6i/YvW6ppqt3KZiqmqJjnExMRO61LgFxEwfE7hnl2pbFVEYvq94zCzTP2HEUxHXj3IneKo9yqAb8AAAAACB3Tq4M3dPahucSNP4aPpNmVyPplbonnh8TMz6fbblRXy96rfs3hFhcXn2U5bnuTYvJ84wVnG5fjLU2sRYu0703KZ7Yn5/B2q5ulJwDzPhXnNWbZPbxGO0ji7n1jETHWqwlU/6q7Mf/wBavD78A4aAAAAAAAAMlpvIc61JmtvKsgyvF5njrnqLGGtTXVPu7R4PdfjnWXX8ozfF5XiqrNWIwl2qzemzci5RFdM7VRFVPKqImJjeJmJ25TMcweNIDoB+yFwvk3Ffuwj+kB0A/ZC4Xybiv3YBYwAAAAACtjp0+yTz74PhP4e24c7j06fZJ598Hwn8PbcOAABMboDcY4s3fqV6hxHpLkzcyO7V2U1c5uWKpmfDymnl29aPDTCaSm7AYvEYDHYfHYO9VZxOHu03bNyntorpnemY96YhZ10YOLGG4rcOrGOxFVFvPsBFOHzSzG0b3Ijldpj+zXHP3J3jwRMh1cAAABwXphcFbPErR853kmGt06qyqia7NURO+LsxEzVYnbtnw0zz5xt91LvQCmmumqiuqiumaaqZ2mJ7Yl8pc9O3gpOWZhe4oaYwNc4PF3Otndm1RHVs3ZmPr+0c4iuZ9NP9qd59UiMAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA3TgtxCzXhlxAwGqcsmqum1V3vF4eJ2jEWKvV0Tv8cT4JiJaWAt90JqnJta6Sy/U+Q4ib+X4+zF23NUbV0T4aKo8FVM7xMeOGbVn9F3jjmfCjUkYPHV3MXpXHV7Y3Cb7zZqnsvW/FVHhjsqj3dpiyHTmdZXqLI8JneS42zjsvxluLti/aq3prpn/wB4neJjtiYmJB7wAAAH8roproqorpiqmqNppmN4mH9ARW4+dEfKtR4rE5/w5vYfJsyvVzcvZdemYwt2qqd5miYiZteGdoiafBEUwhlrjRGrdEZrXlmq9P47KsTT2d+t+krjfbeiuN6a43iedMzHJbs8mbZXlub4KrBZrgMLjsNV6q1iLVNymf0TAKchZhqrov8ABnP4u1f0Yqyq/coimm7luJrs9792KN5t7+/TLScT0KOGlVERhtSaut1eO5iMPXHxRZgEBRPyx0KOGVNva/qPV9de/bRicNTHxTYn9reNNdGPgvkddu7RpKnML1NHUmvMMTcvxX7s0TPU396mAVt5FkedZ9jreByPKMfmeKuT1aLOEw9V2uqdt9oimJnsiZSI4YdDzXmfTaxescVhtMYKrnNnrRfxUx+TTPVp/TVv44hPHJMkyfI8LGFybK8Fl1mIiOphrFNuJiI2jfaOfJr3FziNpvhjpG9qLUeImmiPSYbDW9pu4q7tvFFET4fHPZEc5BxvV+VcKui1w8uZvk2V4bF6rxlM2cuvY/69icRdiOdXg73ap3iaur1Y9TE7zMID53mmPzvOMZnGa4qvFY/G368RiL1fqrlyqZmqqffmW0cZOJGoOKOtMRqTPrkUb/W8JhKKpm3hbO/Kinx+7PhnefcaWAs86HnsdNKfmLnnKlYazzoeex00p+YuecqB1wAAAAAFRnFT1z9VeWcX56trTZeKnrn6q8s4vz1bWgAAAAAAXK2PsFv8mP2Pt8WPsFv8mP2PsAAAAGvcTfW31P5Hxfma1RC3fib62+p/I+L8zWqIAAB1nox8X8dwm1zTibkzeyDMZps5ph9pmepE8rtG33dO8+/EzG3ZMWaZNmWAznKsLmuV4uzjMDi7VN6xfs1RVRcoqjeJiYU5O+9FDj/iuF+ZfSDUFV/F6TxdzeqimetVgq5nncoj+zP3VMdvbHPtCxseTJszy/OcpwubZVjLOMwOLtU3bF+zV1qLlExvExL1gAAAANK4q8LNE8TMqjA6rye3fuW9+8Yy19bxNiZ236lyOe3KN4neJ2jeJ2huoCBnEroZ6yyuu5idD5phNQYaI3pw2Irpw2I8HKJqnqT2z21U9iPeqNEax0vipw2otL5xldyOtt9E4OuimqInaZpqmNqo92JmFu74vWrV+1VZvW6LtuuNqqK6YmKo8UxPaCmoW2Z3w04fZ1ai1mmi8hxNEdkVYG3H7IYf6h3CD8XWnfkdIKq3uyvJ82zW/bsZZleNx127V1LdGHsVXKq6vFEUxO8rV8l4UcNMmuzdyzQuQYaue2acFRP7YbVluXZfltjvGXYHC4O1M79TD2qbdO/vUxAK2NC9GLjDqrvd6vTc5Fg69/r+b3Iw8xtPZ3rnd38W9ERPjSJ4b9DHSOVXbeL1tnOK1Bdoq3+hcPvh8PV28qpievMdnZNPYlO57xw4t6X4UacqzDOsRTezC9TP0Dl1ur67iKo9z7mnftqnl+kGucceImkeAfDeMNkWWZdg8xxFuq1lGV4WzTbo6/Pe5XTTttRTM7zPbMzt2zMxWznma5hnmcYvOM2xdzF4/GXar2IvXPVV11TvM/8AaOUM3xP11qDiJq/F6m1Fipu4m/O1u3T9jsW49TbojwUx/wA+2ectXAABb7oD7RNP+TMN5qlm2E0B9omn/JmG81SzYAAAAMZq37VM3+A3vN1KelwurftUzf4De83Up6AAAAAAAABJHud3r6Y3yHf87ZWDq+O53evpjfId/wA7ZWDgAAAAIa90r+xaN/KxP/QmUhr3Sv7Fo38rE/8AQCGIAAAAAAAAAAAAAAACbHc8OJf0Tl2P4Y5pep75hetjcpqqmmJqtzP1214JmYmevHbO1VfZFMITth4b6szDQ2uMp1VlkzOIy7E03ZtxV1Yu0b+ntzO07RVTvG+07b7+AFu4xWkdQZXqrTGXajyXERiMvzHD04ixXExv1ao7J27KoneJjwTEx4GVAB8Yi9Zw2HuYjEXbdmzaomu5cuVRTTRTEbzMzPKIiPCDkPS14pW+GfC7E14O7tn2bRVg8tpiImaJmPT3p33jaiJ392qaY7N5ismuqquuquuqaqqp3mZneZl1PpQ8T7nFLiji80w01U5Ngd8JlduZnnapnncmPHXO9XuRtHg3nlYAAAAAAAAAACR/c8PX5xXkLEedso4JH9zw9fnFeQsR52yCwkAAAAAFd3dBfZAf4Rhv21o8JD90F9kB/hGG/bWjwAD2ZNlWZ51mFvLsny7GZjjbv2PD4SxVduV+9TTEzIPG+7Nq7eu02rNuu5crnamiineZnxREO68OeipxU1TctXczy+1prA1xFU3cxq2ubTG/K1TvVv7k7c+3ZMPgj0fdCcLqLeMwmHqzfPYj0+aYymJrpnbn3ujstx29m8+OqeQOO9FPowXMrxWB1xxHwsRi6Ipv4DJ7kfYqu2mu/H9qO2KPBPqucbJfgAAAAA17iPrDJtB6NzDVGe34tYPB2+t1d4iq7XPqbdPjqqnlD717rDT2htNYnUOpsxtYHA2I7ap9Ncq8FFFPbVVPgiFbHSK4zZ5xd1TOJv8AfMFkOErmMty7r7xbjs75Xtym5PhnwdkcuchrnGDiBnPEzXmO1VnNXVqv1dTD4emqZow1mnlRbp38Uc5nwzMz4WngAADsHQx9kxo/85if4W8s4Vj9DH2TGj/zmJ/hbyzgAAAABCTulv2xaK+CYv8AftptoSd0t+2LRXwTF/v2wRCAAdd6LPF29wo1/TiMZcuV6fzLq2MztRNUxRTv6W9FMdtVG8+CZ2mqI7XIgFyOXY3B5jgLGPy/FWMXhMRbi7Zv2LkV27lExvFVNUcpiY8MP3QE6HHSCt6Hv0aH1liKo07iLm+DxlVX/wBhcmecVf8Ahz27/czz5xM7T5s3Ld61RetV03LddMVUV0zvFUT2TE+GAfQAAAD8MxwWDzHA3sDmGFsYvC36Zou2b1uK6K6Z8E0zymH7gIa8cuh3VcvYjOeFmIop63WuVZNi7u0b856tm5PKPBERXPv1Ikap0zqHS2aXcs1HkuOyrGWqurVaxVmqiffjflMT4JjeJjnC4FjtQZDkuocDVgc9yrBZlhqomJtYmzTcp5xtO28cv0Ap4Fl+pui7wYzyq9cjTFeWX7sREXMvxVy1FHu00bzbj/haf6CrhZ+H9Z/LMN/8cEABYXk/Q24R4HG038VitT5pbjtsYrHW6aKvfm1boq+Kp0PSnAfhHpmqivK9D5ZN23X16LuKirE10z+VcmqdvcBWzojh5rfW2MowuldL5nmlVVUUzctWZizRvvtNd2raiiJ2nnVMQkzwv6FuPvVWsZxEz6jC2+2rAZZMV3Ozsm7VHVifeieztlNLCYbD4PDW8NhLFrD2LcdWi1aoimmmPFERyhyLpM8csp4R6f7xh4sY/VGNszVgMDVV6WiJ3iL12I597iYnlymqaZiJjnMByvpKav0TwJ0FXw14Z5bhMt1BmuH2xNyzb69yxh6t6ZruXat6qrlcbxTvMzEc/Sx1d4OslqfPs31Nn2Lz3Pcddx2Y4y5Ny/fuTzqmf+UR4IiOUQxoCQHQD9kLhfJuK/dhH9IDoB+yFwvk3FfuwCxgAAAAAFbHTp9knn3wfCfw9tw53Hp0+yTz74PhP4e24cAAA6DwB4m5nwq4iYTUODqquYG5tYzPCxETGIw8zE1RG/ZVExFVMxtzjbsmYnnwC4rIM2wGfZJgs5yrE28TgcbZpvWLtFUVRVTVG8c45Pcg30C+MU5RmtPDDUOO2y/HXN8mru1xFNm/VMzNmJn75M7xG/quURvVznIAAAADzZpgMFmuW4nLcxwtrFYPFWqrN+zdp61FyiqNppmPDExKsjpO8Isbwm15XhLVM3chzDrX8rvxMztRvztVTP3dHLx7xNM+GYi0BpfGfh1k/E/QeN0vm1U2KrsdfCYumiKq8Nej1NcRPbHgmN43iZjeO2AqbGb11pbOtF6qx2ms/wAJVhsfgrk0V0zHKqPBXTPhpmNpifdYQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB17o68dtQ8Is0rs0W6s009ia+tisurudXq1ffLU8+rXt7m07bT4JjkIC3Dhrr/SvETT9GdaVzS1jLO1PfrW8Rdw9Uxv1blHbTPb7k7TtMtoVCaG1hqXRGfWs80tm+JyzHW+XXtTvTXHhprpnemun3JiYTO4M9MTIM3mxlfEfB/STGVTFMZjhqJrwtU8+ddPOq3z2jeOtHPntEbglaPDkecZTnmX28wybMsJmODu0xVRfw16m5RVE9kxNM7PcAAAAAAANe1vrfSWictrzHVef4HKsPTT1vr1z09fbypojequeU8qYmZ2RD41dMbH4+jEZTwzwNeX2Kt6JzXGURN6Y5xvbt84p8cTVvPuR4AkXx4436R4UZTdjH4inHZ7XZmrB5VZq+uXKvuZrn/V0b85qnntE7RM7Qrr4t8SdU8T9TTnmp8b3yqiJowuGt+ls4aiZ36tFPg38MzznlvPKGq5ljsbmePv5hmOLv4zGYi5VdvX79ya7lyuqd5qqqnnMzMzMzLzgAALPOh57HTSn5i55ypWGs86HnsdNKfmLnnKgdcAAAAABUZxU9c/VXlnF+era02Xip65+qvLOL89W1oAAAAAAFytj7Bb/Jj9j7fFj7Bb/Jj9j7AAAABr3E31t9T+R8X5mtUQt34m+tvqfyPi/M1qiAAAAAdi6PXH3VHCjHWsFNVzNdM13Jqv5bcr26m/qqrUz6irw7dkz27b7rB+F3EjSPEnIqc20rmlGJpimmb+Gr2pv4eZj1NyjflPg3jeJ25TKpZldK6iz3Sud2M607muKyzMMPVvbv4evq1R4JifBMTEzExO8TE7SC4MQ34P8ATNs1RYy3ibldVuqZimc1y63vTG87da5a332iJ3maN55cqZ32Sr0frHS2sMuozDTGfYDNcPXG/Ww96Kpp5RO1VPbTMbxvExEwDOgAAAAAAADmPFTjvw14dYe5GcZ/axmYxTvRluXzF/EV9nbET1aI5771zTvETtvPJDfjZ0qNca6sYjKMgidMZJdiaLlGGuTOJvU7zyru8piJjaJinbfnvMxMwCTHSH6S+mOHeGxOS6euWc81Tt1YtUT1sPhat9pm7VHhjafSRz37dkAtdatz/W+psVqLUuY3MdmGJn01dXKminwUU09lNMeCIYOqZqmaqpmZnnMz4X8AAAABb7oD7RNP+TMN5qlm2E0B9omn/JmG81SzYAAAAMZq37VM3+A3vN1KelwurftUzf4De83Up6AAAAAAAABJHud3r6Y3yHf87ZWDq+O53evpjfId/wA7ZWDgAAAAIa90r+xaN/KxP/QmUhr3Sv7Fo38rE/8AQCGIAAAAAAAAAAAAAAAAAJs9zx4lRistzDhnmmK+v4WJxuVRcrneu1M/XbdO8fczMVbb84qq5elmUv1RHDjVeYaI1zlGqssqqjEZdiabvVpqmnvlPZXRO3gqpmaZjxStk0pnmX6l01luoMqvU38DmOGoxNiuN+dNdMTHKdpiee0xMRMT2gyaNXTw4q/0R0La0Vk+JrozrP6Z7/VbnabGDjlXMzE7xVXO1McpiYi52bRvITU2d5ZpvT+Pz7OcVThcuwFiq/iLtUTPVopjedojnM+CIjnM7RCqTi7rjMOInELNdWZh1qJxl6e8WZq63eLMcqLcT7lO2/jneQamAAAAAAAAAAAAkf3PD1+cV5CxHnbKOCR/c8PX5xXkLEedsgsJAAAAABXd3QX2QH+EYb9taPmHvV4fEW79vq9e3XFdPWpiqN4neN4nlMe5PJIPugvsgP8ACMN+2tHgEruAnHThbiqLGQ8UuHek8JiI6tFrObGS2Jt19kfXqIo3pnw9eneOfOKdt5mrpK/p7E5Bhr2lq8tqymunexOAiiLO3uRRyj3lPjbOHfEbW3D7Mfo3SWocblszO9yxTX1rF7lMentVb0VcpnaZjeN+UxILbBCzh/02MRbps4fXWlKb8couYvK6+rVtz3nvVc7T4OXWh3TSHSV4NakotU29X2MqxFdHXqsZpaqw02/cm5VHe5n3Ka5B18YHKNaaQzfC/RWV6pyXG2PvlnHW66fjiXs+n+RfhrLflVHzgyQ0XP8AjDwtyKi/Oaa+0/Zrw8b3LVGMpuXY963RM1VT7kRMuQ656ZXDnKYuWNL5fmuo78bdS73r6Gw87x47np+U7cupHvgkw4zxx6RWhuGeHv4OjEUZ7qCnemjLcJdjair/AMW5zi3EeLnV7nbtDzir0nuJ2ubF7L7GYUadyu7HVrw2WTNFdccuVV31e3KeUTETEzE7uI1TNVU1VTMzM7zM+EG58W+J2reJ+oIzbVGPm5FuJpw2Etb02MNTPbFFG/KZ8MzznaN55Q0sAAAAAdg6GPsmNH/nMT/C3lnCsfoY+yY0f+cxP8LeWcAAAAAISd0t+2LRXwTF/v2020JO6W/bFor4Ji/37YIhAAAAJC9GTpJZtw3vWdPapnFZtpSY6lFNNXWvYHxTb39VR46N427Y7Npj0AuB0hqbIdW5HZzvTmaYbMsBe9Tds17xE7RPVqjtpqiJjeJ582XVLcLuJOsOGud/TXSebXMLVXtF/D1+nsYiI8FyieU+/wApjwTCanBzpdaK1P3nLda2/wCi+Z1RtGIrmasFcnl93225nefVR1do9VziASVH4ZfjcHmGFoxeAxVjFYeuN6Ltm5FdNUe5Mcn7gAAAAAADn3FLjLw84cYSu5qPP7M4yKd7eX4SYvYq7zmOVETy5xMb1TTHLtQv44dKnWWucPicl03RVprI70dS5FmvfFX6OcTFVyPU0zE84p27Nt5iZiQkR0kuk3kWgLF/INI14bO9UdbqXJirrYfBRtvNVcx6uvsiKI8czMxttVAPVGfZvqfUGNz/AD3HXcdmWNud8v37s71VTttHvRERERHZERER2MbMzMzMzMzPbMv4AAAkB0A/ZC4Xybiv3YR/SA6AfshcL5NxX7sAsYAAAAABWx06fZJ598Hwn8PbcOdx6dPsk8++D4T+HtuHAAAAA+7F67h79u/YuV2rtuqK6K6J2qpqid4mJ8ExKzDom8XbHFLh9bozDEUf0lyuIs5jamqOtdj7m/EcvS1eHxVRMeLesxuvBXiHmvDHiBgdUZZNy5RbnveMw0V9WMTYn1VufiiY8UxEgtjGL0lqHKNV6bwGoshxtGMy3H2YvWLtHhifBMdsVRO8TE84mJiecMoAAAACPfTL4KWuIWk6tT5BhKI1TlVuavSxO+Mw8bzVanbtqjtpn348PKuyYmJmJjaY7YXLoK9OngrXkWb3eJemsDcqyvML2+b2rNuOrhb9X+tmI7Ka57Z7OtPOd6oBFEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGd0hrHVOkcbGM0zn+Y5Te8M4a/VRFXZ2x2T2R2w7/AKH6Z3EHKbFGG1Pk+U6joopmO/xE4XEVTv21TRvbmIjltFEe+jGAn5p7po8OsZTV9OMkz3K6opjbq26L0VT4dtqobplnSi4J4zDReu6uqwVU/wCrxGBv9aP+GiqP+as8BZ36Jfgf7fMP8hxP+Wxmd9KzgtltNM2dQ4vM9/Bg8BdmY9/rxSrXATr1T02NI4WL1GndK5rmVdO3eq8TcpsW6vHvt1qo+JxXiH0tuKmpou4bJ7+C0vgq+tT1cvt9a/VTPZFV2veYmP7VEUSj8A9udZtmmd4+vH5xmOKzDF1+qvYm7Vcrnnv2z7sz8bxAAAAAAkHwu6VOrdAaGy3SWXabyPF4XAUVUUXb83evVvVNXPaqI8KPgCVPo29de1HTfx3v5z0beuvajpv47386KwCVPo29de1HTfx3v5z0beuvajpv47386KwCVPo29de1HTfx3v5z0beuvajpv47386KwD36jzS7neoMxzm/botXcfiruJroo36tNVdc1TEb+DeXgAAAAAAAEp6OmzrqmimmNI6c5Rt23v5316NvXXtR038d7+dFYBKn0beuvajpv473856NvXXtR038d7+dFYBKn0beuvajpv473856NvXXtR038d7+dFYBJvUHTJ1rnOQ5hlF/SunrdrHYW5hq66JvdammumaZmN6+3mjIAAAAAAAD3ZJm+a5Jj6Mfk+Y4vL8VbmKqb2Gu1W6omJ3jnE+N4QHetA9LHi1pmKbGY4/BamwkRTTFGZ2d7lMRPPa7bmmqap8dfX952XTvTdyK9RTTqDROYYS5Ne01YPFU3qKafH6aKZ39zZCABY5lnS44N4y9Fu7mGbYKJjnXiMBV1Y/4Zmf8Akz0dJfgfMb/08w8f/scT/lqxQFmuM6T3BHD2KrtGtKcRMfcWsDiOtPx0RH/NqWc9MnhZhcLXXl+Dz7ML0TtFuMNTair3etNSvcBMLVvTezGrvtvSeiMJZ7O938zxFVzbx727fV/fcK1/x54ra2qu0Zvq3GWMJcmf9DwO2Gs00zO8U7UbTVt4JqmqfdcyAf2qZqqmqqZmZneZnwv4AAAAAAAJO5L0zNbZXk+Cyy1pTT1dvCYe3Yoqqm9vVFFMUxM+n7eT2ejb117UdN/He/nRWASp9G3rr2o6b+O9/Oejb117UdN/He/nRWASp9G3rr2o6b+O9/Oejb117UdN/He/nRWAShzPpoa3x+W4rA3NJ6dpoxFmu1VVTN7eIqiYmY9P7qLwAAAAAAAAA3rglxNzjhRq+7qXJMBgcbibmErwk28ZFc0RTVVTVM+lqid/SR4XavRs8Qvatpb/AIL/APmItgJSejZ4he1bS3/Bf/zD0bPEL2raW/4L/wDmItgJSejZ4he1bS3/AAX/APMPRs8Qvatpb/gv/wCYi2AlJ6NniF7VtLf8F/8AzHLuPHG7UXGCnK6c9yrKsB9LZrm19BU3I63X23369VXi8DlgAAAAAAAAAAAAAAAAAAAm13O3iLOJyrMuGmPvTVXg+tj8t3607Wqqo77RHPaIiuqKtuXOuqefghKzugNVZtojWeV6ryO5TRmGW34u2uvEzTXG0xVRVEbTNNVM1UztMTtVPOASw7oRxTp/0bhblF+rrxNGLziqnlERMb2rPZ4piudp/s9u87QyZLVGd5hqTUWYZ/mt7v2Ox+IrxF+vx1VTvO3uMaAAAAAAAAAAAAAkf3PD1+cV5CxHnbKOCR/c8PX5xXkLEedsgsJAAAAABXd3QX2QH+EYb9taPCQ/dBfZAf4Rhv21o8AAAAAAAAAAAAAAAAA2jhVrPHcPdf5XrHLcJh8XisuquVW7OI63e6uvbqtzv1ZieyuZ/QkD6NvXXtR038d7+dFYBKn0beuvajpv473856NvXXtR038d7+dFYBKn0beuvajpv473856NvXXtR038d7+dFYBKn0beuvajpv47387knHzjLnfGHHZTi86yrL8vqyy1ct24wk17VxXNMzv1pn+zDmIAAAAAAAADZtE6/wBaaKxMX9LalzLKpiYmaLN6e91bTv6aid6ao38ExMS79orpp65y7D28NqnT2U57FFMU/RNmasLfrnfnVXt1qJnbwU00Qi2An5kHTQ4c4vrRm2S59lkxEbTFui9FU/oqjZt2UdKngrj7cV3dS4jL5mPU4rA3d4/4KalaoCzi90muCFu3NUa4s3Jj7mnA4nefjttazHpg8IsNF2MPXneMroj0ne8F1aa59yaqo2/TCuwBNDVHTfsxb6mmNC11VzRP1zMcXERRV4J6tET1o9zrQ4dxD6SPFvWk3bWJ1HOUYG5HVnBZTR9D29pp2mJq3m5VE+KquY8WzkAD7v3bt+9XevXK7tyuetVXXVM1VT45me18AAAAAA3XgvxFzLhdre3qvKcBhMdibeHuWItYnrdSYrjaZ9LMTu0oBKn0beuvajpv473856NvXXtR038d7+dFYBKn0beuvajpv473856NvXXtR038d7+dFYBKn0beuvajpv473856NvXXtR038d7+dFYBt/GDXuYcS9e43V+aYLC4LFYui1RVZw3W73TFFEURt1pmeylqAAAAAAAA69wY6Q2veFWQYjIsitZTmGAu3u/UWsytXbkWKpjn1Opco2ie2Y5829+jV4p/gDRnyPE//IRmASZ9GrxT/AGjPkeJ/wDkHo1eKf4A0Z8jxP8A8hGYBJn0avFP8AaM+R4n/wCQejV4p/gDRnyPE/8AyEZgEmfRq8U/wBoz5Hif/kPHnfTC4iZ1lGLyjNdK6IxeBxlqqziLFzBYmablFUbTE/6QjkA+7tVNd2uum3TapqqmYop32pjxRvMzt78y+AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB9W6K7lXUt0VV1T4KY3kHyN20Zwl4l6xqtTp3ROdYuzeiZt4mrDzZw9W3b9eudW3v7nWdx0J0LtZ5jNF7V2fZfklmdpm1h98TeiJj9FMTE8u2ffBFqimquumiimaqqp2ppiN5mfEz+q9Gah0rl2V4zUOAry6rNbVV7CYe9yvVWonbvlVHbTTM9nW2me2I25rIOFPAHhlw2ppxmX5RTj8ztx1qszzKYu3adoneaN4im3HOfUxE7dszsgJ0kdd18ROMWeahpmuMDTd+hMBRVVv1MPa9LTt4utMVVzHgmuQc5AAAAAAAAAAAAAASP7nh6/OK8hYjztlHBI/ueHr84ryFiPO2QWEgAAAAAru7oL7ID/CMN+2tHhIfugvsgP8ACMN+2tHgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB3Loi8INNcXc/z3Aakx2b4S1l+Ft3rU5fdt0VVVVVzE9br0V7xy8GzhqW3c1ftx1f5PseckHSfQVcLPw/rP5Zhv/jnoKuFn4f1n8sw3/x0mQEasP0LeFNq7Fdecavv0x20V42xET/w2In/AJtgw3RN4J2rUUXMgx9+qPu7mZ3omf8AhqiP+TuoDlmU9Hjg1luGixa0Jlt+I+7xPWu1/wDFVMy3fJNH6TySu1XlGmsnwNyzT1bdyxg7dFdMe5VEb/8ANnAAAHF+mTxBo0HwYx9vDYmm3m+d75fgaYn00RVH125Eb77U0b8+yKqqIntVnu3dMriXTxC4s4jD5diu/wCSZH1sFg5ormaLlUT9duR4OdUbb+GKY5zGziIAAAAAAAAAAAAAACR/c8PX5xXkLEedso4JC9APMcvyzjhicTmWOwuCszkmIpi5iLtNunrTcs7RvVMRvykFiYwf9MdI+2nI/wBYWv5j+mOkfbTkf6wtfzAzgwf9MdI+2nI/1ha/mP6Y6R9tOR/rC1/MDODB/wBMdI+2nI/1ha/mP6Y6R9tOR/rC1/MCBXdBfZAf4Rhv21o8O/dPPMMBmfHf6Jy3HYbG2PpTh6e+Ye7Tcp3ia943pmY3cBAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAS27mr9uOr/J9jzkokpbdzV+3HV/k+x5yQTiAAAAAAcD6afFi1w/4dV5DluKpp1Dn9quxYppn09jDzG1y7PKYjt6sb7TMzMx6mZjrHEvWmScP9GY/VOf4im1hMJR6WjeIrv3J9Taojw1VTy+OZ5RMqteK2uc34i65zDVec1bXsVX9asxXNVOHtR6i3Tv4Ij453nwg1UAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABLbuav246v8n2POSiSlt3NX7cdX+T7HnJBOIAAABhda6pyLRunMTqDUeYWcDgMPG9Vdyraap8FNMfdVTtyiGp8aOM2iuFmVV388x9OIzKqP9HyvDVxViL0zvtMx9xRy51VcvfnaJry448X9VcWs9tY7PblGHwOF3jB5fYme82Intnn6qqeW9U+KAe/pHcZM44uatnE3OvhMiwdVVOW4HrcqKfvlfjrq8M+Dsj3eVgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlt3NX7cdX+T7HnJRJblwv4mav4a4vHYrSOYW8Fex1qm1fqrsUXN6aZ3jbrRO3OQWzPm7ct2qJru100UxG8zVO0RCrvHdIXjLi7tddevcztRXG002upRTHvREcmjag1VqfUFMU57qHNczoiua4oxWLru001T4YiqZiP0Asm170iuEekKKqMRqzCZrio//AE2VVRiqu2Y51UT1KZiY5xNUT7iMPF3ph6t1Bh72V6GwFOmsHciaasbXVF3GVUzG3pZ9Tb7Z5xE1dkxNMwi8A9OaZhjs0x97H5ljL+Mxd+qa7t6/cmuuuZ7ZmZ5y8wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA6ppLo/cVNVacwWock09Ticux1vvmHu/RNFPWp3mN9pneOcSynoXuNHtWp+V2/nTf6J/sddF/AJ85W6iCs/wBC9xo9q1Pyu3856F7jR7Vqfldv51mACs/0L3Gj2rU/K7fznoXuNHtWp+V2/nWYAKz/AEL3Gj2rU/K7fznoXuNHtWp+V2/nWYAKz/QvcaPatT8rt/Oehe40e1an5Xb+dZgArP8AQvcaPatT8rt/Oehe40e1an5Xb+dZgArP9C9xo9q1Pyu3856F7jR7Vqfldv51mACrzVfR84q6X07jc/zrT1OHy/A2pu4i79E0VdWmPDtE7y5UtL6VHse9Z+Ta/wBsKtAAAAAG9cM+EuvOI2CxeN0lks47D4S5Fq9cm7TREVzG+3OefL9rRqaaqqoppiaqpnaIiOcytO6NOgfqc8H8myLEWu95ndtRi8xiYjenEXIiqqidpnfqcqN99p6u/hBBj0L3Gj2rU/K7fznoXuNHtWp+V2/nWYAKz/QvcaPatT8rt/O0/iZwp1zw4w+CxGrsn+gbWNqqosVRdpriqqmImY5Ty7YWwOZdJ3QNPEXg3nOS2rc1Zjh7f0dl0x2/RFqJmKecxHpo61HPs6+/gBVoPqumqiuqiumaaqZ2qpmNpifE+QAAAAAAHQeHHBviBxCye9m+lMmjG4OzemzXXN+ijauIidtpnxTDnyfvc6PWgzbyvX5ugEavQvcaPatT8rt/Oehe40e1an5Xb+dZgAqo4k8GuIPDzI7OdasyWMDgb+Jpwtu5F+ive5VTVVEbRPioq+Jz5P8A7o76yGTf3ksfw2JQAAAAAAAAAAAAbtwy4V624kW8dc0jlUY+nATRTiJm9TR1Jr63V7Z5+plpKanc0P6u1x+ewX7LwOJ+he40e1an5Xb+c9C9xo9q1Pyu386zABWf6F7jR7Vqfldv5z0L3Gj2rU/K7fzrMAFZ/oXuNHtWp+V2/nPQvcaPatT8rt/OswAVn+he40e1an5Xb+c9C9xo9q1Pyu386zABWf6F7jR7Vqfldv5z0L3Gj2rU/K7fzrMAFZ/oXuNHtWp+V2/nPQvcaPatT8rt/OswAVI8SuH+qeHeb4fKdWZfGBxeIw8Yi1RFymve3NVVO+8T46ZaqlF3SH129P8AkGj+IvIugAA6lo7gDxR1dprB6iyLT9OJy3G0zXYu/RFFPWiKppnlM79sSy3oXuNHtWp+V2/nTX6H/scNH/BrvnrjrIKz/QvcaPatT8rt/OwGvuBvEzQ+nrmf6i07Xh8utV00XL1F2m51Jq5RMxTO8Rvy392Fp7yZ1lmX51lOKynNcHZxmBxdqq1fsXad6blExtMTAKch2HpRcGcdwo1nXVg7V69pjMK5qy3FVem6nhmzXP8Aap8HjjafHtx4AAAAAAB2j0L3Gj2rU/K7fzuMR2wuXBWf6F7jR7Vqfldv5z0L3Gj2rU/K7fzrMAFPeq8gzTS+osbp/OsP9D5hgbk2r9rrRV1avFvHJi3TulV7IXWflGr92HMQHWNN9HfixqHIcFnmVacpv4HHWab+HufRNuOtRVG8TtM8nJ1rXRz9YrRfkex+6CBvoXuNHtWp+V2/nPQvcaPatT8rt/OswAVn+he40e1an5Xb+c9C9xo9q1Pyu386zABWf6F7jR7Vqfldv5z0L3Gj2rU/K7fzrMAFZ/oXuNHtWp+V2/nPQvcaPatT8rt/OswAVn+he40e1an5Xb+c9C9xo9q1Pyu386zABWf6F7jR7Vqfldv5z0L3Gj2rU/K7fzrMAFZ/oXuNHtWp+V2/nYDX3A7iToXTd3UOpcipweXWq6Ldd2MRRVtVVO1MbRO/atPcG6efsc80+HYTzsArgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABaP0T/Y66L+AT5yt1FFPgB0i+FmkuDmmtOZ3m+Ks5jgMJ3q/RThK6opq69U9sdvKYb16K3gx+HcZ8hrB3IcN9FbwY/DuM+Q1noreDH4dxnyGsHchw30VvBj8O4z5DW92nukxwlz7P8uyPLc6xdzG5jirWEw1FWDrpiq5criimJnwc5jmDsgAAOP6o6SXCjTeoswyDNs5xVrH5fiK8PiKKcHXVFNdM7TETHbzB2AcN9FbwY/DuM+Q1noreDH4dxnyGsHchw30VvBj8O4z5DWeit4Mfh3GfIawbL0qPY96z8m1/thVonfx56RvCvVfCDUuncmzjFXswx+CqtWKKsJXTFVUzHhnsQQAAAAB2noaaD/pxxry6vE2O+ZbksfTHFb8o3omO909sdtc08vFE+CJWXuB9BrQE6N4O282xmHm1meorkY29NVG1UWYjazRPh2iJqq9+5LvgAAAAKy+mHw/r0FxozGnD2IoyrOf/qOAmmmYpiK5nvlvxb01xVy5+lmmZ7XG1i3Ts0D/AEs4P3M+wlmq5mGm6qsZHVp3mcPMbXo7N9oiIrns9Rv4FdIAAAAAACfvc6PWgzbyvX5uhAJP3udHrQZt5Xr83QCTYAIzd0d9ZDJv7yWP4bEoAJ/90d9ZDJv7yWP4bEoAAAAAAAAAAAAJqdzQ/q7XH57BfsvIVpqdzQ/q7XH57BfsvAmIAANW4qa5yfhzozE6rz2jE14HD3LduuLFHWr3rqimNo38cuMejJ4Wf7Jn3yWn+YEkBG/0ZPCz/ZM++S0/zHoyeFn+yZ98lp/mBJARv9GTws/2TPvktP8AMejJ4Wf7Jn3yWn+YEkBG/wBGTws/2TPvktP8x6MnhZ/smffJaf5gSQEb/Rk8LP8AZM++S0/zHoyeFn+yZ98lp/mBxnukPrt6f8g0fxF5F12npd8UNPcVdd5VnenLWMt4bC5ZThbkYm3FFXXi7cq5REzy2qhxYAAFn/Q/9jho/wCDXfPXHWXJuh/7HDR/wa75646yAADXOJWi8j4gaNx+ltQYfvuDxdG0VU+rs1xzpuUT4KqZ5/HE7xMwq94y8PM64Y67xumM5t1zFuevhMT1JijFWZ9Tcp8finxTEx4FsrmvSH4T5ZxZ0Jdye9VawubYfe7luNqo37zd/s1eHqVdk7e5PPbYFWIyurtPZtpXUmO09nmFqwuYYG7Nq9bnnzjwxPhie2J8UsUAAAAD+x2wuXU0R2wuXAABVr0qvZC6z8o1fuw5i6d0qvZC6z8o1fuw5iAta6OfrFaL8j2P3VUqfPBvpJ8KNN8K9M5Dmuc4q1jsBl1qxiKKcHXVFNdNO0xvHaCT44b6K3gx+HcZ8hrPRW8GPw7jPkNYO5DhvoreDH4dxnyGs9FbwY/DuM+Q1g7kOVaB6QPDLXGrMHpfT2bYnEZnjO+d5t14WuiJ6lFVdXOeUelpl1UAAAcXzjpO8IcpzfGZVjc7xdGKwd+vD3qYwVcxFdFU01Rv4ecS8voreDH4dxnyGsHchw30VvBj8O4z5DWeit4Mfh3GfIawdycG6efsc80+HYTzsP39FbwY/DuM+Q1uUdK/j1w319wax2m9NZricRmN7FYe5RRXhaqImmiuJq5z7gIYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAANz4E+vfoP+8mXfxNtpjc+BPr36D/ALyZd/E2wWygAKp+kf6/euPLeJ85K1hVP0j/AF+9ceW8T5yQc/AAAAAAAAbzwI0Le4jcVMk0tTFUYW/fi5jrlM7Tbw1HprkxO07TNMdWOW3Wqjdoycnc7OH9WA0xmnEPMMPNN3M7k4PLpqpjnYtz9cuRPimvenwfY58cAlfg8NYweEs4TC2aLOHsW6bdq3RG1NFNMbRTEeCIiNn6gAjjY6Q03elfXw6i7hf6NdX6XUXerHWqx0RvNXX5cpq+t7c45RMc5da44a4w3Dvhbnmq7+1V3C4eaMJb++Yiv0tqntjl1piZ25xTFU+BVRGa5lTncZ3Rjb9GZRifoqMVRX1bkXut1uvEx2VdbnvHhBcWNO4La1scQuGOSartd7i9jMNH0VbomNrd+n0tynaJnaOtEzETz2mG4g/LGYbD4zCXsHi7FrEYa/bqt3bVymKqLlFUbVU1RPKYmJmJiVUPHDRVzh7xSzzStUVd4wuImrC1VdtVir01ufd9LMRv44Wxokd0U4fTjtP5TxFwFmZv5dV9A5ht4bFczNuuef3Ne9PKJme+R4KQQeAAAAAAT97nR60GbeV6/N0IBJ+9zo9aDNvK9fm6ASbABGbujvrIZN/eSx/DYlABP/ujvrIZN/eSx/DYlAAAAAAAAAAAABNTuaH9Xa4/PYL9l5CtNTuaH9Xa4/PYL9l4ExAAcL6dvsb86+FYTz9KttZJ07fY3518Kwnn6VbYAAAAAAAAAAAALP8Aof8AscNH/BrvnrjrLk3Q/wDY4aP+DXfPXHWQAfnXfs0X6LFd63TduRM0UTVEVVbdu0eHYH6AAj/0vuBmH4kabr1Hp7A0Rq7L7e9E0elnG2o7bVXgmqPuZ7fB2K68TYvYbE3cNibNyzftVzRct3KZpqoqidppmJ5xMTy2XKIe9ODgP9F2sZxR0lh6qsTRtXnODt0798p7PoiiI8McutHi9N4J3CFIAAAP7HbC5dTRHbC5cAAFWvSq9kLrPyjV+7DmLp3Sq9kLrPyjV+7DmIAAAAAAO0dCP2TelPexv8HfWYKz+hH7JvSnvY3+DvrMAAAVGcVvXR1Z5bxnn62tNl4reujqzy3jPP1taAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAbnwJ9e/Qf95Mu/ibbTG58CfXv0H/eTLv4m2C2UABVP0j/X71x5bxPnJWsKp+kf6/euPLeJ85IOfgAAAAAAAyukMjxmptU5Xp7LqOti8yxVvDWo/wB6uqI/91tmi8gwWltJ5VpzLrdNvC5dhbeHtxG/ZTG2/PnzneefjQv7nboCMw1bmXELHWJmzlVucJgJqjl3+5TtXVE/7tuZp/8A8k+JOcAGP1LnGC09p7MM9zG53vB4DDV4i9Vv2U0UzM/sBDPui2vacXneUcO8DiKaqcDTGOzCmivfq3a42t0VRE9sUTNW0xvtXTPhRDZ3iBqXHax1rm+qMyuVV4nMsVXfq3+5iZ9LTHOdopp2piPBERDBAmB3OniDRh80zbhvmF3aMXE4/LJqntrpja7bjn4adq4iI+5r38CbCofhxqjGaL1zk+qMBVVF7LsVRe6tM7demJ9NT+mneP0ra8hzPCZ1keAznAXIu4TH4a3ibFcdlVu5TFVM/piYB7WF13pzBau0dm2mcwiPobMsLXh6pmN+pNUcqtt432nae3wM0Ap61ZkeYaZ1Pmens1td6x2W4q5hr9Pg61FUxMx44nbeJ8MTEsYlX3RDQX0q1pluvcHYqjD5zRGGxlURyjEW6fSzPLlNVEeGefUnxIqAAAAAJ+9zo9aDNvK9fm6EAk/e50etBm3levzdAJNgAjN3R31kMm/vJY/hsSgAn/3R31kMm/vJY/hsSgAAAAAAAAAAAAmp3ND+rtcfnsF+y8hWmp3ND+rtcfnsF+y8CYgAOF9O32N+dfCsJ5+lW2tw4p6GybiNo3E6Uz+vFUYDEXLdyucNciiveiqKo2mYnwx4nGfQbcJ/9q1H8so/kBXqLCvQbcJ/9q1H8so/kPQbcJ/9q1H8so/kBXqLCvQbcJ/9q1H8so/kPQbcJ/8AatR/LKP5AV6iwr0G3Cf/AGrUfyyj+Q9Btwn/ANq1H8so/kBXqLCvQbcJ/wDatR/LKP5ECNV4Gzlmp80y3DzVNnC4y7ZtzVO89WmuYjf3doBjAAAAWf8AQ/8AY4aP+DXfPXHWXJuh/wCxw0f8Gu+euOsgIidPvU+caO1zw21HkOKnD4/A/Rt23V201bVWN6ao8NMxvEx4pS7Qr7ph/WOhPzOO/bYBJLgHxTyjixoa1n+AtxhcbanvWYYKa+tOHu7dm/hpntieXJ0JVBwR4lZ3wt1zhdQ5TduVYfrRRj8HFW1GKs786Jjs38MT4J/StA4f6vyLXWk8FqbTuMpxOAxdG9M9lVuqOVVFcfc1RPKYBn38rpproqorpiqmqNpiY3iY8T+gK+emZwKr0Jnd7W2msPH9GMxv/XbFEf8A2F6r7n83VO/V8W/V8W8bVxWoMoyzP8lxeS5zgrONy/GWptYixdp3prpn/wD7eJ8E7TCs7pMcHcz4T60u2qLF67pzG3Kq8rxk+miae3vVU+Cun3e2OcA5MAD+x2wuXU0R2wuXAABVr0qvZC6z8o1fuw5i6d0qvZC6z8o1fuw5iAAAAAADtHQj9k3pT3sb/B31mCs/oR+yb0p72N/g76zAAAFRnFb10dWeW8Z5+trTZeK3ro6s8t4zz9bWgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAG58CfXv0H/eTLv4m20xufAn179B/wB5Mu/ibYLZQAFU/SP9fvXHlvE+clawqn6R/r9648t4nzkg5+AAAAAA/TDWL2JxNrDYe1XdvXa4ot0URvVVVM7RER4ZmX5u89B/QFes+MtjNMVhpuZVp6iMbiKpp3om7vtZtzvExvNUTVt4rdQJycA9B2OG/CrJdL0xbqxduzF7H3KOy5ia/TXJido3iJ9LG8RPVpjdvYAIvd0J4g15Hw/wOhsvv9TGZ9c75i5pq2qpwtuYnq/+evqx71FUT2pQzMREzMxER2zKrHpLa+p4j8YM5z/C3ZuZZbufQuXT1ZjrYe3ypr2mImOtO9XON46209gOagAJ89z111GdcN8ZorF3oqxeQ3prw8TPOcPdqmrxeCvreHwx4IQGdO6MOv7nDrjHk+c3K5jLsTX9A5jTEb9axcmImeyZ9LV1a+XOept2TILSh/KZiqmKqZiYmN4mPC/oNA6QehMPxF4S53pu5Zt14ubM4jAV1U7zaxNv01FVPKZiZ50ztzmmuqPCqpv2rli/csXqKrdy3VNFdNUbTTMTtMSuVVudNvh7OieMuKzLCWopynUMTj8Nt2UXZna9b/RX6aOW21cRHZIOFAAAAJ+9zo9aDNvK9fm6EAk/e50etBm3levzdAJNgAjN3R31kMm/vJY/hsSgAn/3R31kMm/vJY/hsSgAAAAAAAAAAAAmp3ND+rtcfnsF+y8hWmp3ND+rtcfnsF+y8CYgAAAAAAAAACoTiF9vmf8AlLEecqW9qhOIX2+Z/wCUsR5yoGCAAABZ/wBD/wBjho/4Nd89cdZcm6H/ALHDR/wa75646yAhX3TD+sdCfmcd+2wmohX3TD+sdCfmcd+2wCHbtPRV4143hTqynB5hdru6VzK9TGYWdpq7xM8u/wBEeOI23iPVRG3bEOLALkMsx2DzPL8PmGX4q1isJibcXLN61VFVFymY3iYmO2HoQM6FPHidK5hY4dapvb5Jjb+2XYqur/7O7VO3Unf/AFdU7fkzM+CeU8wGrcVNC5HxG0XjdLZ/bqnDYiN7d2jbvli5HqblO/hifj7G0gKlOLegc64ba4x2ls6p61zD1b2MRFE00Ym1PqblO/gmPBz2nePA1JaP0keEWW8WtD15fNNjD55g97uWY2unnbr8NFUxz6lXZMeOIntiFZOpckzTTmf47Is6wdzB5jgb1VjEWa4501UztPvx4YmOUxMTHKQY+O2Fy6miO2Fy4AAKtelV7IXWflGr92HMXTulV7IXWflGr92HMQAAAAAAdo6Efsm9Ke9jf4O+swVn9CP2TelPexv8HfWYAAAqM4reujqzy3jPP1tabLxW9dHVnlvGefra0AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA3PgT69+g/7yZd/E22mNz4E+vfoP+8mXfxNsFsoACqfpH+v3rjy3ifOStYVT9I/1+9ceW8T5yQc/AAAAAAWV9C3QcaK4KYDE4izTRmOezGY4irw9SqI71TPvUbT/wCafHKDXRv0D9Ubi7k2n79qqvLqbv0TmO3+z2/TVU9sbdblTy5x1t/AtSpiKaYppiIiI2iI8AAAMFxByPG6l0Vm+n8vzecoxGY4WvDU42LHfZsxVG1UxR1qd+W8dsdqKPoGKfxoT+of/wAhMsBDT0DFP40J/UP/AOQegYp/GhP6h/8AyEywENPQMU/jQn9Q/wD5B6Bin8aE/qH/APITLAYXQuT43T2jspyPMc1+m2JwGFow9eN7z3qb3VjaKpo61W07RHhlmgAcM6bOgatacF8bj8HaivMtPzOY2eW812qYnv1Mc/7G9Xh9TtEc3c38rpproqoqjemqNpj3AU0DoXSI0LPDvi5nenLdqq3gqbv0Rgd+ybFz01G0+HbnT/5XPQAAE/e50etBm3levzdCASfvc6PWgzbyvX5ugEmwARm7o76yGTf3ksfw2JQAT/7o76yGTf3ksfw2JQAAAAAAAAAAAATU7mh/V2uPz2C/ZeQrTU7mh/V2uPz2C/ZeBMQAHH+mHqLPNK8CM1znTuZ4nLMwtYnDU28Rh6urXTFV2mJiJ92J2QP+rvxh/GJn3yj/ALJudO32N+dfCsJ5+lW2DpH1d+MP4xM++Uf9j6u/GH8YmffKP+zm4DpH1d+MP4xM++Uf9j6u/GH8YmffKP8As5uA6R9XfjD+MTPvlH/Y+rvxh/GJn3yj/s5uA6R9XfjD+MTPvlH/AGc8xmIv4zF3cXibtV2/erm5crqnnVVM7zM/pfkAAAAAs/6H/scNH/BrvnrjrLk3Q/8AY4aP+DXfPXHWQEK+6Yf1joT8zjv22E1EK+6Yf1joT8zjv22AQ7AATo6E/HqjPMDY4dayzOqc4sx1MqxWIq3nFW4jlamqe2unwb85jlzmEF375fjMVl+OsY7BX7mHxOHuU3bN2idqqKoneJiffBciOH9E3jbZ4q6VqwGcVWLGqctpinFW6J2jE2+yL9NPg37Ko8E8+yYiO4AI69MbgVRxCyGvVemMFT/SvAUb10Ucpx9mI9RPjrj7mffjwxtIoBTVdt3LN+qzet127lFU010VxtVTMTtMTE9krlUPOmP0dsXmWZ1a+4f5bfxWNxd6IzTLbFHWmuqf9fbjxz91H6fHvMMAAFWvSq9kLrPyjV+7DmLp3Sq9kLrPyjV+7DmIAAAAAAO0dCP2TelPexv8HfWYKz+hH7JvSnvY3+DvrMAAAVGcVvXR1Z5bxnn62tNl4reujqzy3jPP1taAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAbnwJ9e/Qf95Mu/ibbTG58CfXv0H/AHky7+JtgtlAAVT9I/1+9ceW8T5yVrCqfpH+v3rjy3ifOSDn4AAAAM9w80xmGtNb5PpbK7VdzFZliqLMdWJnqU9tdc7RO1NNMVVTPgimZBNfue3D36SaCxuvMfa6uNz2vvWE3id6MLbnbf8A89e8+9RTMTzSjY/TWT4LT+n8vyPLrfe8HgMPRh7NO0R6WmmIjs5b8mQABE3ppcetS6G1blmktB51ZwWNs2PonNLlNm3eqomv7HanrxVFM9X08xtE7VUT2SCWQrN9FDxw9un/AKdhv8s9FDxw9un/AKdhv8sFmQrN9FDxw9un/p2G/wAs9FDxw9un/p2G/wAsFmQrN9FDxw9un/p2G/yz0UPHD26f+nYb/LBZkKzfRQ8cPbp/6dhv8tILoV8etS661Tmukdd5rTj8ddsRisrvfQ9u1M9TfvtuepERM7TTVHL7mvn2AlgACLPdDdAU5xobLteYHDROOyW53jF10Ub1V4W5PLrTEb7UV843mIjvlfjQNXDaqyTAal03mOn80tzcwWYYavD3ojbfq1RMbxvExEx2xPgmIVK6+0xmOi9aZtpXNop+jMsxNViuqn1NcR6mun/dqpmKo9yYBgwAE/e50etBm3levzdCASfvc6PWgzbyvX5ugEmwARm7o76yGTf3ksfw2JQAT/7o76yGTf3ksfw2JQAAAAAAAAAAAATU7mh/V2uPz2C/ZeQrTU7mh/V2uPz2C/ZeBMQAHC+nb7G/OvhWE8/SrbWSdO32N+dfCsJ5+lW2AAAAAAAAAAAACz/of+xw0f8ABrvnrjrLk3Q/9jho/wCDXfPXHWQEK+6Yf1joT8zjv22E1EK+6Yf1joT8zjv22AQ7AAABmdE6mzjR2qsv1LkOMuYXMMBei7arpqmIq8dFXjpqjeJjsmJmFn/AfijkvFfQ9rPssmLGLtT3nMMFVV6fDXtt5j3aZ7aavDHuxMRVO3zgdxNzzhXrjDZ/lVyu5haqot5hg+ttRirO/Omf96O2mfBPubwC1wYLQWrMi1vpXBal07jaMXl+Lo61NUeqoq+6oqj7mqJ5TDOgAAAAq16VXshdZ+Uav3YcxdO6VXshdZ+Uav3YcxAAAAAAB2joR+yb0p72N/g76zBWf0I/ZN6U97G/wd9ZgAACozit66OrPLeM8/W1psvFb10dWeW8Z5+trQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADc+BPr36D/ALyZd/E22mNz4E+vfoP+8mXfxNsFsoACqfpH+v3rjy3ifOStYVT9I/1+9ceW8T5yQc/AAAATA7nRoKMRmuccRMbYpqowkTl+Aqqjfa5VETdqj3Ypmmnft2qmPCiPl2DxOY5hh8vwVqb2JxN2m1Ztx21V1TtEc/dlbFwb0XhOH3DPI9JYSN/oLDR3+vw3L9U9e7X+muqraPBG0eAG3AA8Wf5pg8kyPHZzmFyLWEwOHrxF6uZ26tFFM1TPxQqU4j6oxmtNd5zqnH11VX8yxdd703bTT2UU9s7RTTFNMR4IhN7ugmv6sg4a4XRmAv8AUxuoLv8ApPVq2qpwtuYmqOXP01fVjxTEVxPagCAAAAAAA2ThjqzGaH19k2qsFNc3MvxVN2uimdpuW99q6P00zMfpa2AuNybMcHm+T4LNsvvU38HjcPbxGHu09lduumKqao9yYmJetGnufmuoz/hdiNIYq91sZp69taie2cPcmaqfB4KuvHb4vcSWAQh7otoH6CzvJ+ImCtbWcfH0BjpinlF6mJqt1Tz+6oiqOz7jt5wm807jToyxr/hfnulL1HWrxmGmcPO+3VvUTFdud947K6ae3l4+QKmB+2Nw1/B4y9g8Tbm3fsXKrdyie2mqmdpj44fiAn73Oj1oM28r1+boQCT97nR60GbeV6/N0Ak2ADz47BYLH2Ys47CYfFW4q60UXrcV0xPZvtPh5z8bx/0b07+Acq+R2/mZQBi/6N6d/AOVfI7fzH9G9O/gHKvkdv5mUAYv+jenfwDlXyO38x/RvTv4Byr5Hb+ZlAGL/o3p38A5V8jt/Mf0b07+Acq+R2/mZQBi/wCjenfwDlXyO38x/RvTv4Byr5Hb+ZlAEOu6O5ZluX6Z0fVgMuwmEqrxmJiqbNmmiao6lHbtHNCxN7uln2r6M+G4n9yhCEBNTuaH9Xa4/PYL9l5CtNTuaH9Xa4/PYL9l4ExAAY/UOSZRqHK7mV55luFzHA3Jpqrw+ItxXRVMTvEzE+KebV/qQcLvaDp35DR8zeAGj/Ug4Xe0HTvyGj5j6kHC72g6d+Q0fM3gBo/1IOF3tB078ho+Y+pBwu9oOnfkNHzN4AaP9SDhd7QdO/IaPmPqQcLvaDp35DR8zeAGj/Ug4Xe0HTvyGj5j6kHC72g6d+Q0fM3gBFLpwaA0TpvgpOY5BpXKcsxn0ysUd+w2Gpor6s9beN4jsQUWI90D9YOfKmH/AOpXcAACz/of+xw0f8Gu+euOsuTdD/2OGj/g13z1x1kBCvumH9Y6E/M479thNRCvumH9Y6E/M479tgEOwAAAAAds6KfG3G8KtVRgMwqnEaWzO7TGOsz22KuyL9Hux4Y8Me7ETFk2AxeGx+BsY7B37eIw2It03bN23V1qa6Ko3pqiY7YmJiVNqU/Qm48WtK42zw71djblOSYu7tlmKuVb0YK7VO/e6v7Nuuqe2OVNU7ztE1TATxCJiYiYneJ7JAAAVa9Kr2Qus/KNX7sOYundKr2Qus/KNX7sOYgAAAAAA7R0I/ZN6U97G/wd9ZgrP6Efsm9Ke9jf4O+swAABUZxW9dHVnlvGefra02Xit66OrPLeM8/W1oAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABufAqYjjdoOZ5R/STLv4m20x9W667dym5brqorpmJpqpnaYmOyYkFynfLf3yn4zvlv75T8anf6cZt+FMd8oq+c+nGbfhTHfKKvnBcR3y398p+NVT0jpiePWuJid4+neJ85LTfpxm34Ux3yir53ju3K7tyq5drqrrqneqqqd5mfHMg+QAAASI6BugadU8WZ1Lj7dM5Zpu3GI9P2XMTXvFqntjs2qr8PqKYmPTLDu+W/vlPxqccLjsbhaJowuMxFimZ3mLdyaYmf0S/b6cZt+FMd8oq+cFxHfLf3yn4zvlv75T8anf6cZt+FMd8oq+d/JzfNpjaczxsxP/j1fODovSl1/TxF4y5xm2Eu98yvCV/QOXTHZVZtzMdePy6utV49piJ7HLQAAAAAAAAB1bopa+u8PuNOT5hcvTRlmYVfS/Madt4qs3JiInsmfS1xRXy5z1ZjsmVoHfLf3yn41NL2xm+axERGZ42IjsiL9XzguJ75b++U/Gd8t/fKfjU7/TjNvwpjvlFXzn04zb8KY75RV84O79O7QNGluLteo8vsRRlmo6PoqepT6WjExtF6N+zeqdrnjma6vEj09GKxuMxVNNOKxeIvxTO8RcuTVt8bzgJ99zqrpp4Q5tFVUR/9Xr7Z/wDDoQEenDY/HYWiaMNjMRYomd5pt3aqY3/RILje+W/vlPxnfLf3yn41O/04zb8KY75RV859OM2/CmO+UVfOC4jvlv75T8Z3y398p+NTv9OM2/CmO+UVfOfTjNvwpjvlFXzguI75b++U/Gd8t/fKfjU7/TjNvwpjvlFXzn04zb8KY75RV84LiO+W/vlPxnfLf3yn41O/04zb8KY75RV859OM2/CmO+UVfOC4jvlv75T8Z3y398p+NTv9OM2/CmO+UVfOfTjNvwpjvlFXzguI75b++U/Gd8t/fKfjU7/TjNvwpjvlFXzn04zb8KY75RV84Jn90qqpq0xo3q1RP+m4nsn/AHKEInoxWNxmKimMVi79+KedMXLk1be9u84Cafc0qqacu1x1qoj69gu2fcvIWP3wuMxmEiqMLir9jreq73cmnf39gXH98t/fKfjO+W/vlPxqd/pxm34Ux3yir5z6cZt+FMd8oq+cFxHfLf3yn4zvlv75T8anf6cZt+FMd8oq+c+nGbfhTHfKKvnBcR3y398p+M75b++U/Gp3+nGbfhTHfKKvnPpxm34Ux3yir5wXEd8t/fKfjO+W/vlPxqd/pxm34Ux3yir5z6cZt+FMd8oq+cFxHfLf3yn4zvlv75T8anf6cZt+FMd8oq+c+nGbfhTHfKKvnBcR3y398p+M75b++U/Gp3+nGbfhTHfKKvnPpxm34Ux3yir5wWA90Croq4CTFNVMz9NMP2T+UrwenE5hj8Tb73iMbib1G+/VuXaqo396ZeYAAFnvRArojo46Pia6Yn6GueH/AMa46z3y398p+NTpZzLMbFqm1ZzDF2rdPqaaL1URH6Il9/TjNvwpjvlFXzguI75b++U/GhZ3S+qmrMdC9WqJ+s47sn3bCJX04zb8KY75RV878MVi8Xi5pnFYq/f6vqe+XJq297cH4AAAAAAP7EzExMTMTHZMP4And0J+PNepsDZ4eaxx1uc5wtvbK8Xdq2qxlqmN+91TPbcpiOU9tVMc95iZmVHfLf3yn41Nduuu1cpuW66qK6Z3pqpnaYnxxL2fTjNvwpjvlFXzguI75b++U/Gd8t/fKfjU7/TjNvwpjvlFXzn04zb8KY75RV84OgdKmYnpCazmJ3j6Y1fuw5i+7125euVXb1yu5XVO9VVU7zPvy+AAAAAAAdn6EkxHSb0pMzERtjO34HfWX98t/fKfjU22L17D3Yu2Ltdq5T2V0VTTMfph6vpxm34Ux3yir5wXEd8t/fKfjO+W/vlPxqd/pxm34Ux3yir5z6cZt+FMd8oq+cGX4rc+KOrJj8NYzz9bWn9rqqrrmuuqaqqp3mZneZl/AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAf/Z"
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>CPT | Campaign Reporter</title>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@700;800;900&family=DM+Sans:wght@300;400;500&display=swap');
    *,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
    :root{{--bg:#f5f4ef;--card:#ffffff;--border:rgba(0,0,0,0.10);--olive:#4a5a0a;--black:#111111;--muted:rgba(17,17,17,0.50);--cream:#f5f4ef;--grad:linear-gradient(135deg,#4a5a0a 0%,#8a9e20 100%)}}
    html,body{{min-height:100vh;background:var(--bg);color:var(--black);font-family:'DM Sans',sans-serif;font-size:15px}}
    #top-bar{{position:fixed;top:0;left:0;right:0;height:60px;background:#111;display:flex;align-items:center;justify-content:space-between;padding:0 36px;z-index:100;border-bottom:3px solid var(--olive)}}
    .top-bar-logo img{{height:26px}}
    .top-bar-badge{{font-family:'Poppins',sans-serif;font-size:11px;font-weight:700;letter-spacing:2.5px;text-transform:uppercase;color:#c8d840;background:rgba(200,216,64,0.12);padding:5px 12px;border-radius:100px;border:1px solid rgba(200,216,64,0.3)}}
    .page{{position:relative;z-index:1;min-height:100vh;padding:88px 24px 60px;display:flex;flex-direction:column;align-items:center}}
    .hero{{text-align:center;margin-bottom:36px;animation:fadeUp 0.6s ease both}}
    .hero-eyebrow{{display:inline-block;font-size:11px;font-weight:600;letter-spacing:3px;text-transform:uppercase;color:var(--olive);margin-bottom:14px;padding:5px 14px;border:1px solid rgba(74,90,10,0.3);border-radius:100px;background:rgba(74,90,10,0.06)}}
    .hero h1{{font-family:'Poppins',sans-serif;font-size:clamp(30px,5vw,50px);font-weight:900;line-height:1.1;letter-spacing:-1.5px;margin-bottom:14px}}
    .hero h1 em{{font-style:normal;background:var(--grad);-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
    .hero p{{font-size:15px;color:var(--muted);max-width:460px;margin:0 auto}}
    .card{{background:var(--card);border:1px solid var(--border);border-radius:20px;padding:36px;width:100%;max-width:640px;box-shadow:0 4px 32px rgba(0,0,0,0.08);animation:fadeUp 0.6s ease both;position:relative;overflow:hidden}}
    .card::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:var(--grad)}}
    .platform-label{{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:1.5px;margin-bottom:10px;font-weight:600}}
    .platform-toggles{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px}}
    .toggle-chip{{display:flex;align-items:center;gap:7px;padding:8px 16px;border-radius:100px;font-size:13px;font-weight:500;border:1.5px solid var(--border);color:var(--muted);background:rgba(0,0,0,0.02);cursor:pointer;transition:all 0.2s;user-select:none}}
    .toggle-chip input{{display:none}}
    .toggle-chip.active{{border-color:var(--olive);background:rgba(74,90,10,0.08);color:var(--black);font-weight:600}}
    .platform-section{{margin-bottom:24px}}
    .drop-zone{{border:2px dashed rgba(0,0,0,0.15);border-radius:14px;padding:48px 24px;text-align:center;cursor:pointer;transition:all 0.25s;background:rgba(0,0,0,0.02);position:relative}}
    .drop-zone:hover,.drop-zone.drag-over{{border-color:var(--olive);background:rgba(74,90,10,0.05)}}
    .drop-zone input[type=file]{{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}}
    .drop-icon{{font-size:36px;margin-bottom:12px}}
    .drop-text{{color:var(--muted);font-size:14px}}
    .drop-text strong{{color:var(--black);display:block;margin-bottom:4px;font-size:15px;font-weight:600}}
    #file-selected{{display:none;align-items:center;gap:12px;padding:14px 18px;background:rgba(74,90,10,0.06);border:1.5px solid rgba(74,90,10,0.25);border-radius:12px}}
    #file-selected .file-name{{flex:1;font-size:14px;font-weight:500}}
    .remove-btn{{background:none;border:none;color:var(--muted);cursor:pointer;font-size:18px;padding:2px 6px;border-radius:6px}}
    .remove-btn:hover{{color:#c0392b}}
    .btn{{width:100%;padding:16px;border:none;border-radius:12px;font-family:'DM Sans',sans-serif;font-size:15px;font-weight:600;cursor:pointer;transition:all 0.25s;display:flex;align-items:center;justify-content:center;gap:8px;margin-top:12px}}
    .btn-primary{{background:var(--black);color:white;box-shadow:0 4px 16px rgba(0,0,0,0.2)}}
    .btn-primary:hover:not(:disabled){{background:var(--olive);transform:translateY(-2px)}}
    .btn-primary:disabled{{opacity:0.35;cursor:not-allowed;transform:none}}
    .btn-outline{{background:transparent;border:1.5px solid var(--border);color:var(--muted)}}
    .btn-outline:hover{{border-color:var(--olive);color:var(--black)}}
    #progress-screen,#complete-screen{{display:none}}
    .progress-header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}}
    .progress-label{{font-size:14px;color:var(--muted)}}
    .progress-pct{{font-family:'Poppins',sans-serif;font-size:14px;font-weight:700;color:var(--olive)}}
    .progress-track{{height:5px;background:rgba(0,0,0,0.08);border-radius:100px;overflow:hidden;margin-bottom:24px}}
    .progress-fill{{height:100%;background:var(--grad);border-radius:100px;width:0%;transition:width 0.5s ease}}
    #log-box{{background:#111;border-radius:12px;padding:16px;height:300px;overflow-y:auto;font-family:'Courier New',monospace;font-size:12px;line-height:1.7;color:rgba(255,255,255,0.55)}}
    #log-box span{{display:block}}
    .complete-icon{{width:64px;height:64px;background:var(--grad);border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:28px;margin:0 auto 20px;box-shadow:0 4px 24px rgba(74,90,10,0.35)}}
    .complete-title{{font-family:'Poppins',sans-serif;font-size:26px;font-weight:800;text-align:center;margin-bottom:10px}}
    .complete-sub{{text-align:center;color:var(--muted);font-size:14px;margin-bottom:20px}}
    .template-links{{text-align:center;margin-top:14px;font-size:12px;color:var(--muted)}}
    .template-links a{{color:var(--muted);text-decoration:none}}
    .template-links a:hover{{color:var(--olive)}}
    .spinner{{width:16px;height:16px;border:2px solid rgba(255,255,255,0.3);border-top-color:white;border-radius:50%;display:inline-block;animation:spin 0.7s linear infinite}}
    @keyframes spin{{to{{transform:rotate(360deg)}}}}
    @keyframes fadeUp{{from{{opacity:0;transform:translateY(20px)}}to{{opacity:1;transform:translateY(0)}}}}
    #log-box::-webkit-scrollbar{{width:4px}}
    #log-box::-webkit-scrollbar-thumb{{background:rgba(255,255,255,0.15);border-radius:2px}}
  </style>
</head>
<body>
  <div id="top-bar">
    <div class="top-bar-logo"><img src="data:image/png;base64,{LOGO}" alt="Cherry Pick Talent"></div>
    <div class="top-bar-badge">Campaign Reporter</div>
  </div>
  <div class="page">
    <div id="main-screen">
      <div class="hero">
        <div class="hero-eyebrow">End of Campaign</div>
        <h1>Campaign <em>Reporter</em></h1>
        <p>Upload your campaign spreadsheet and we'll auto-fill views, engagement, CCV and Rebrandly click data — ready for your EOC deck in minutes.</p>
      </div>
      <div class="card">
        <div class="platform-section">
          <div class="platform-label">Select platforms to pull data for:</div>
          <div class="platform-toggles">
            <div class="toggle-chip active" id="toggle-yt" onclick="togglePlatform('yt')"><input type="checkbox" id="check-yt" checked><span>🎬 YouTube</span></div>
            <div class="toggle-chip active" id="toggle-tw" onclick="togglePlatform('tw')"><input type="checkbox" id="check-tw" checked><span>🟣 Twitch</span></div>
            <div class="toggle-chip" id="toggle-ig" onclick="togglePlatform('ig')"><input type="checkbox" id="check-ig"><span>📸 Instagram</span></div>
            <div class="toggle-chip" id="toggle-tt" onclick="togglePlatform('tt')"><input type="checkbox" id="check-tt"><span>🎵 TikTok</span></div>
            <div class="toggle-chip active" id="toggle-rb" onclick="togglePlatform('rb')"><input type="checkbox" id="check-rb" checked><span>🔗 Rebrandly</span></div>
          </div>
          <div style="font-size:12px;color:var(--muted);opacity:0.7;">💡 Instagram & TikTok require Modash API — coming soon</div>
        </div>
        <div class="drop-zone" id="drop-zone" ondragover="onDragOver(event)" ondragleave="onDragLeave(event)" ondrop="onDrop(event)">
          <input type="file" accept=".xlsx" id="file-input" onchange="onFileSelect(event)">
          <div class="drop-icon">📊</div>
          <div class="drop-text"><strong>Drop your campaign report here</strong>or click to browse — .xlsx only</div>
        </div>
        <div id="file-selected">
          <span style="font-size:20px;">📄</span>
          <span class="file-name" id="file-name-label"></span>
          <button class="remove-btn" onclick="removeFile()">✕</button>
        </div>
        <button class="btn btn-primary" onclick="runJob()" id="run-btn" disabled>
          <span id="run-btn-text">Pull Campaign Data</span>
        </button>
        <div class="template-links">
          <a href="https://docs.google.com/spreadsheets/d/1nQDKHV9GcWCWXVDyY_-wSNUOQCHPXV9au172VLIvI-Y/copy" target="_blank">📊 Open template in Google Sheets</a>
        </div>
      </div>
    </div>

    <div id="progress-screen">
      <div class="hero">
        <div class="hero-eyebrow">Pulling Data</div>
        <h1>Building Your <em>Report</em></h1>
      </div>
      <div class="card">
        <div class="progress-header">
          <span class="progress-label" id="progress-label">Fetching stats...</span>
          <span class="progress-pct" id="progress-pct">0%</span>
        </div>
        <div class="progress-track"><div class="progress-fill" id="progress-bar"></div></div>
        <div id="log-box"></div>
      </div>
    </div>

    <div id="complete-screen">
      <div class="hero">
        <div class="hero-eyebrow">Complete</div>
        <h1>Report <em>Ready</em></h1>
      </div>
      <div class="card">
        <div class="complete-icon">✅</div>
        <div class="complete-title">All done!</div>
        <div class="complete-sub">Your campaign data has been pulled and your EOC deck is ready.</div>
        <button class="btn btn-primary" onclick="downloadPPTX()">🎞️ &nbsp;Download EOC PowerPoint Deck</button>
        <button class="btn btn-outline" onclick="downloadFile()">📊 &nbsp;Download Filled Excel Report</button>
        <button class="btn btn-outline" onclick="startOver()">Run Another Report</button>
      </div>
    </div>
  </div>

  <script>
  let currentJobId = null, selectedFile = null, pollInterval = null;

  function togglePlatform(p) {{
    var chip = document.getElementById('toggle-' + p);
    var check = document.getElementById('check-' + p);
    if (!chip || !check) return;
    check.checked = !check.checked;
    chip.classList.toggle('active', check.checked);
  }}
  function onDragOver(e) {{ e.preventDefault(); document.getElementById('drop-zone').classList.add('drag-over'); }}
  function onDragLeave(e) {{ document.getElementById('drop-zone').classList.remove('drag-over'); }}
  function onDrop(e) {{ e.preventDefault(); document.getElementById('drop-zone').classList.remove('drag-over'); if (e.dataTransfer.files[0]) setFile(e.dataTransfer.files[0]); }}
  function onFileSelect(e) {{ if (e.target.files[0]) setFile(e.target.files[0]); }}
  function setFile(file) {{
    if (!file.name.endsWith('.xlsx')) {{ alert('Please upload an .xlsx file.'); return; }}
    selectedFile = file;
    document.getElementById('file-name-label').textContent = file.name;
    document.getElementById('file-selected').style.display = 'flex';
    document.getElementById('drop-zone').style.display = 'none';
    document.getElementById('run-btn').disabled = false;
  }}
  function removeFile() {{
    selectedFile = null;
    document.getElementById('file-selected').style.display = 'none';
    document.getElementById('drop-zone').style.display = 'block';
    document.getElementById('run-btn').disabled = true;
    document.getElementById('file-input').value = '';
  }}
  async function runJob() {{
    if (!selectedFile) return;
    document.getElementById('run-btn-text').innerHTML = '<span class="spinner"></span>&nbsp; Starting...';
    document.getElementById('run-btn').disabled = true;
    try {{
      var form = new FormData();
      form.append('file', selectedFile);
      form.append('run_yt', document.getElementById('check-yt').checked ? '1' : '0');
      form.append('run_tw', document.getElementById('check-tw').checked ? '1' : '0');
      form.append('run_ig', document.getElementById('check-ig').checked ? '1' : '0');
      form.append('run_tt', document.getElementById('check-tt').checked ? '1' : '0');
      form.append('run_rb', document.getElementById('check-rb').checked ? '1' : '0');
      var res = await fetch('/api/run', {{method: 'POST', body: form}});
      var data = await res.json();
      if (data.job_id) {{
        currentJobId = data.job_id;
        document.getElementById('main-screen').style.display = 'none';
        document.getElementById('progress-screen').style.display = 'block';
        if (pollInterval) clearInterval(pollInterval);
        pollInterval = setInterval(pollStatus, 2000);
      }} else {{ alert(data.error || 'Something went wrong.'); resetBtn(); }}
    }} catch(e) {{ alert('Network error: ' + e.message); resetBtn(); }}
  }}
  function resetBtn() {{
    document.getElementById('run-btn-text').textContent = 'Pull Campaign Data';
    document.getElementById('run-btn').disabled = false;
  }}
  async function pollStatus() {{
    if (!currentJobId) return;
    try {{
      var res = await fetch('/api/status/' + currentJobId);
      if (!res.ok) return;
      var data = await res.json();
      document.getElementById('progress-bar').style.width = (data.progress||0) + '%';
      document.getElementById('progress-pct').textContent = (data.progress||0) + '%';
      document.getElementById('progress-label').textContent =
        data.status === 'complete' ? 'Complete!' : data.status === 'error' ? 'Error occurred' : 'Fetching stats...';
      var box = document.getElementById('log-box');
      if (data.log && data.log.length > 0) {{
        box.innerHTML = data.log.map(function(l){{ return '<span>' + l.replace(/</g,'&lt;') + '</span>'; }}).join('');
        box.scrollTop = box.scrollHeight;
      }}
      if (data.status === 'complete') {{
        clearInterval(pollInterval); pollInterval = null;
        setTimeout(function(){{
          document.getElementById('progress-screen').style.display = 'none';
          document.getElementById('complete-screen').style.display = 'block';
        }}, 800);
      }} else if (data.status === 'error') {{
        clearInterval(pollInterval); pollInterval = null;
      }}
    }} catch(e) {{ console.warn('Poll error:', e); }}
  }}
  async function downloadFile() {{
    var res = await fetch('/api/download/' + currentJobId);
    if (!res.ok) {{ alert('Download failed.'); return; }}
    var blob = await res.blob();
    var a = document.createElement('a'); a.href = window.URL.createObjectURL(blob);
    a.download = 'CPT_EOC_Report_Filled.xlsx'; document.body.appendChild(a); a.click(); document.body.removeChild(a);
  }}
  async function downloadPPTX() {{
    var res = await fetch('/api/download-pptx/' + currentJobId);
    if (!res.ok) {{ alert('PowerPoint not available.'); return; }}
    var blob = await res.blob();
    var a = document.createElement('a'); a.href = window.URL.createObjectURL(blob);
    a.download = 'CPT_EOC_Deck.pptx'; document.body.appendChild(a); a.click(); document.body.removeChild(a);
  }}
  function startOver() {{
    currentJobId = null; selectedFile = null;
    if (pollInterval) clearInterval(pollInterval);
    document.getElementById('log-box').innerHTML = '';
    document.getElementById('progress-bar').style.width = '0%';
    document.getElementById('progress-pct').textContent = '0%';
    resetBtn(); removeFile();
    document.getElementById('main-screen').style.display = 'block';
    document.getElementById('progress-screen').style.display = 'none';
    document.getElementById('complete-screen').style.display = 'none';
  }}
  </script>
</body>
</html>"""

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000, debug=False)
